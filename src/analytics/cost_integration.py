"""Cost-Schedule Integration — CBS/WBS correlation and budget analysis.

Parses cost breakdown structure (CBS) data from Excel files and
correlates with schedule WBS for integrated cost-schedule views.

Supports:
- CBS hierarchy parsing (Level 1, Level 2, Scope, Design Package)
- WBS element budget totals
- CBS-to-WBS mapping extraction
- Budget variance analysis (original vs current)

Reference: AACE RP 10S-90 — Cost Engineering Terminology;
           PMI Practice Standard for EVM — CBS/WBS integration.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from typing import Any

logger = logging.getLogger(__name__)


@dataclass
class CBSElement:
    """A single Cost Breakdown Structure element."""

    cbs_code: str
    cbs_level1: str = ""
    cbs_level2: str = ""
    scope: str = ""
    design_package: str = ""
    wbs_code: str = ""  # Mapped WBS element code
    estimate: float = 0.0
    contingency: float = 0.0
    escalation: float = 0.0
    budget: float = 0.0


@dataclass
class WBSBudget:
    """Budget totals for a WBS element."""

    wbs_code: str
    wbs_name: str = ""
    total_budget: float = 0.0
    cost: float = 0.0
    contingency: float = 0.0


@dataclass
class CBSWBSMapping:
    """Mapping between CBS cost categories and WBS elements."""

    cost_category: str
    cbs_code: str
    wbs_level1: str = ""
    wbs_level2: str = ""
    notes: str = ""


@dataclass
class CostIntegrationResult:
    """Complete cost-schedule integration analysis."""

    cbs_elements: list[CBSElement] = field(default_factory=list)
    wbs_budgets: list[WBSBudget] = field(default_factory=list)
    cbs_wbs_mappings: list[CBSWBSMapping] = field(default_factory=list)
    total_budget: float = 0.0
    total_contingency: float = 0.0
    total_escalation: float = 0.0
    program_total: float = 0.0
    budget_date: str = ""
    insights: list[str] = field(default_factory=list)


def parse_cbs_excel(file_path: str) -> CostIntegrationResult:
    """Parse CBS data from an Excel file.

    Reads the 'Summary by CBS', 'Summary by WBSE', and mapping sheets
    from a standard program budget workbook.

    Args:
        file_path: Path to the Excel file (.xlsx).

    Returns:
        CostIntegrationResult with parsed CBS, WBS budgets, and mappings.

    Reference: AACE RP 10S-90 — Cost Engineering Terminology.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — CBS parsing unavailable")
        return CostIntegrationResult(insights=["openpyxl required for CBS parsing"])

    result = CostIntegrationResult()

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        logger.error("Failed to open CBS Excel: %s", exc)
        return CostIntegrationResult(insights=[f"Failed to open file: {exc}"])

    # Parse "Summary by CBS" sheet
    if "Summary by CBS" in wb.sheetnames:
        ws = wb["Summary by CBS"]
        _parse_cbs_summary(ws, result)

    # Parse "Summary by WBSE" sheet
    if "Summary by WBSE" in wb.sheetnames:
        ws = wb["Summary by WBSE"]
        _parse_wbs_summary(ws, result)

    # Parse CBS-WBS mapping (Sheet1 / Appendix B)
    for sheet_name in ["Sheet1", "Mapping", "CBS-WBS"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _parse_cbs_wbs_mapping(ws, result)
            break

    # Parse budget date
    if "Summary by CBS" in wb.sheetnames:
        ws = wb["Summary by CBS"]
        for row in ws.iter_rows(min_row=1, max_row=3, max_col=5, values_only=True):
            for cell in row:
                if cell and hasattr(cell, "strftime"):
                    result.budget_date = cell.strftime("%Y-%m-%d")
                    break

    # Generate insights
    _generate_insights(result)

    return result


def _parse_cbs_summary(
    ws: Any,
    result: CostIntegrationResult,
) -> None:
    """Parse CBS hierarchy from 'Summary by CBS' sheet."""
    header_row = None
    headers: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
        values = [c.value for c in row]
        # Find header row
        if any(str(v).startswith("CBS") for v in values if v):
            header_row = row_idx
            for i, v in enumerate(values):
                if v:
                    headers[str(v).strip()] = i
            continue

        if header_row is None or row_idx <= header_row:
            continue

        # Parse data rows
        def get_val(col_name: str, default: str = "") -> str:
            idx = headers.get(col_name)
            if idx is not None and idx < len(values) and values[idx]:
                return str(values[idx]).strip()
            return default

        def get_num(col_name: str) -> float:
            idx = headers.get(col_name)
            if idx is not None and idx < len(values) and values[idx]:
                try:
                    return float(values[idx])
                except (ValueError, TypeError):
                    return 0.0
            return 0.0

        lvl1 = get_val("CBS-Lvl1")
        lvl2 = get_val("CBS-Lvl2")
        scope = get_val("Scope")
        wbs_code = get_val("WBSE L1")

        if not lvl1 and not lvl2 and not scope:
            continue
        if lvl1 == "Total":
            result.total_budget = get_num("DJV Estimate")
            result.total_contingency = get_num("DJV Contingency (25%)")
            result.total_escalation = get_num("Escalation")
            continue

        cbs_code = f"{lvl1}.{lvl2}" if lvl2 else lvl1
        elem = CBSElement(
            cbs_code=cbs_code,
            cbs_level1=lvl1,
            cbs_level2=lvl2,
            scope=scope,
            design_package=get_val("Design Package"),
            wbs_code=wbs_code,
            estimate=get_num("DJV Estimate"),
            contingency=get_num("DJV Contingency (25%)"),
            escalation=get_num("Escalation"),
        )
        elem.budget = elem.estimate + elem.contingency + elem.escalation
        result.cbs_elements.append(elem)


def _parse_wbs_summary(
    ws: Any,
    result: CostIntegrationResult,
) -> None:
    """Parse WBS budget totals from 'Summary by WBSE' sheet."""
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip headers
        if row[0] and row[1]:
            try:
                wbs = WBSBudget(
                    wbs_code=str(row[0]).strip(),
                    total_budget=float(row[1]) if row[1] else 0.0,
                )
                result.wbs_budgets.append(wbs)
                result.program_total += wbs.total_budget
            except (ValueError, TypeError):
                continue


def _parse_cbs_wbs_mapping(
    ws: Any,
    result: CostIntegrationResult,
) -> None:
    """Parse CBS-to-WBS mapping from mapping sheet."""
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Look for rows with CBS code pattern (C.XX.XXXXXX)
        values = [str(v).strip() if v else "" for v in row]
        for i, val in enumerate(values):
            if val.startswith("C.") and "." in val[2:]:
                mapping = CBSWBSMapping(
                    cost_category=values[i - 1] if i > 0 else "",
                    cbs_code=val,
                    wbs_level1=values[i + 1] if i + 1 < len(values) else "",
                    notes=values[i + 2] if i + 2 < len(values) else "",
                )
                result.cbs_wbs_mappings.append(mapping)


def _generate_insights(result: CostIntegrationResult) -> None:
    """Generate insights from cost data."""
    if result.cbs_elements:
        total = sum(e.estimate for e in result.cbs_elements)
        construction = sum(
            e.estimate for e in result.cbs_elements if "Construction" in e.cbs_level1
        )
        if total > 0:
            result.insights.append(
                f"Total estimate: ${total / 1e9:.2f}B "
                f"(construction: ${construction / 1e9:.2f}B = "
                f"{construction / total * 100:.0f}%)"
            )

    if result.wbs_budgets:
        top = sorted(result.wbs_budgets, key=lambda w: w.total_budget, reverse=True)[:3]
        for w in top:
            result.insights.append(f"WBS {w.wbs_code}: ${w.total_budget / 1e6:.0f}M budget")

    if result.total_contingency > 0 and result.total_budget > 0:
        pct = result.total_contingency / result.total_budget * 100
        result.insights.append(f"Contingency: {pct:.0f}% of estimate")

    if result.cbs_wbs_mappings:
        result.insights.append(
            f"CBS-WBS mappings: {len(result.cbs_wbs_mappings)} cost categories mapped"
        )


@dataclass
class CBSElementDelta:
    """Per-CBS-element variance between two snapshots.

    Positive ``budget_delta`` means the element's budget grew from A to B;
    ``variance_pct`` is the signed percent change (``(b - a) / a * 100``)
    with zero-safe handling when A is zero.
    """

    cbs_code: str
    cbs_level1: str = ""
    cbs_level2: str = ""
    scope: str = ""
    budget_a: float = 0.0
    budget_b: float = 0.0
    estimate_a: float = 0.0
    estimate_b: float = 0.0
    contingency_a: float = 0.0
    contingency_b: float = 0.0
    budget_delta: float = 0.0
    estimate_delta: float = 0.0
    contingency_delta: float = 0.0
    variance_pct: float = 0.0
    status: str = "changed"  # 'changed', 'unchanged', 'added', 'removed'


@dataclass
class CostComparisonResult:
    """Diff between two CBS cost snapshots.

    Compares two ``CostIntegrationResult`` instances by ``cbs_code`` to
    surface budget variance per element plus program-level totals. Used
    for trend/variance analysis across forecast revisions.

    Reference: AACE RP 10S-90 — Cost Engineering Terminology;
               AACE RP 29R-03 §5.3 — variance documentation.
    """

    snapshot_a: str = ""
    snapshot_b: str = ""
    total_budget_a: float = 0.0
    total_budget_b: float = 0.0
    total_budget_delta: float = 0.0
    total_contingency_a: float = 0.0
    total_contingency_b: float = 0.0
    total_contingency_delta: float = 0.0
    total_escalation_a: float = 0.0
    total_escalation_b: float = 0.0
    total_escalation_delta: float = 0.0
    program_total_a: float = 0.0
    program_total_b: float = 0.0
    program_total_delta: float = 0.0
    budget_variance_pct: float = 0.0
    element_deltas: list[CBSElementDelta] = field(default_factory=list)
    added_count: int = 0
    removed_count: int = 0
    changed_count: int = 0
    unchanged_count: int = 0
    insights: list[str] = field(default_factory=list)


def compare_cost_snapshots(
    a: CostIntegrationResult,
    b: CostIntegrationResult,
    snapshot_a_id: str = "",
    snapshot_b_id: str = "",
) -> CostComparisonResult:
    """Compute element-level and program-level variance between two snapshots.

    Matches CBS elements by ``cbs_code``. Elements present only in A are
    flagged ``removed``; only in B are ``added``. Budget deltas are
    signed (B minus A) so positive values indicate growth.

    Args:
        a: Earlier snapshot (baseline).
        b: Later snapshot (current).
        snapshot_a_id: Optional identifier echoed in the result.
        snapshot_b_id: Optional identifier echoed in the result.

    Returns:
        ``CostComparisonResult`` with per-element deltas, totals,
        counts by status, and interpretation insights.

    Reference: AACE RP 10S-90 (Cost Engineering Terminology);
               AACE RP 29R-03 §5.3 (Forensic Schedule Documentation).
    """
    result = CostComparisonResult(
        snapshot_a=snapshot_a_id,
        snapshot_b=snapshot_b_id,
        total_budget_a=a.total_budget,
        total_budget_b=b.total_budget,
        total_budget_delta=b.total_budget - a.total_budget,
        total_contingency_a=a.total_contingency,
        total_contingency_b=b.total_contingency,
        total_contingency_delta=b.total_contingency - a.total_contingency,
        total_escalation_a=a.total_escalation,
        total_escalation_b=b.total_escalation,
        total_escalation_delta=b.total_escalation - a.total_escalation,
        program_total_a=a.program_total,
        program_total_b=b.program_total,
        program_total_delta=b.program_total - a.program_total,
    )

    if a.total_budget > 0:
        result.budget_variance_pct = (b.total_budget - a.total_budget) / a.total_budget * 100

    a_by_code = {e.cbs_code: e for e in a.cbs_elements if e.cbs_code}
    b_by_code = {e.cbs_code: e for e in b.cbs_elements if e.cbs_code}

    for code in sorted(set(a_by_code) | set(b_by_code)):
        ea = a_by_code.get(code)
        eb = b_by_code.get(code)
        if ea is not None and eb is not None:
            budget_delta = eb.budget - ea.budget
            variance_pct = (budget_delta / ea.budget * 100) if ea.budget > 0 else 0.0
            status = "unchanged" if abs(budget_delta) < 0.01 else "changed"
            delta = CBSElementDelta(
                cbs_code=code,
                cbs_level1=eb.cbs_level1 or ea.cbs_level1,
                cbs_level2=eb.cbs_level2 or ea.cbs_level2,
                scope=eb.scope or ea.scope,
                budget_a=ea.budget,
                budget_b=eb.budget,
                estimate_a=ea.estimate,
                estimate_b=eb.estimate,
                contingency_a=ea.contingency,
                contingency_b=eb.contingency,
                budget_delta=budget_delta,
                estimate_delta=eb.estimate - ea.estimate,
                contingency_delta=eb.contingency - ea.contingency,
                variance_pct=variance_pct,
                status=status,
            )
            if status == "changed":
                result.changed_count += 1
            else:
                result.unchanged_count += 1
        elif eb is not None:
            delta = CBSElementDelta(
                cbs_code=code,
                cbs_level1=eb.cbs_level1,
                cbs_level2=eb.cbs_level2,
                scope=eb.scope,
                budget_b=eb.budget,
                estimate_b=eb.estimate,
                contingency_b=eb.contingency,
                budget_delta=eb.budget,
                estimate_delta=eb.estimate,
                contingency_delta=eb.contingency,
                variance_pct=100.0 if eb.budget > 0 else 0.0,
                status="added",
            )
            result.added_count += 1
        else:
            assert ea is not None
            delta = CBSElementDelta(
                cbs_code=code,
                cbs_level1=ea.cbs_level1,
                cbs_level2=ea.cbs_level2,
                scope=ea.scope,
                budget_a=ea.budget,
                estimate_a=ea.estimate,
                contingency_a=ea.contingency,
                budget_delta=-ea.budget,
                estimate_delta=-ea.estimate,
                contingency_delta=-ea.contingency,
                variance_pct=-100.0 if ea.budget > 0 else 0.0,
                status="removed",
            )
            result.removed_count += 1
        result.element_deltas.append(delta)

    _generate_compare_insights(result)
    return result


def _generate_compare_insights(result: CostComparisonResult) -> None:
    """Build human-readable interpretation of a cost comparison."""
    direction = "increase" if result.total_budget_delta > 0 else "decrease"
    if abs(result.total_budget_delta) >= 1.0:
        result.insights.append(
            f"Program budget {direction}: "
            f"${abs(result.total_budget_delta) / 1e6:.1f}M "
            f"({result.budget_variance_pct:+.1f}%)"
        )

    if result.added_count:
        result.insights.append(f"{result.added_count} new CBS elements added in snapshot B")
    if result.removed_count:
        result.insights.append(f"{result.removed_count} CBS elements removed from snapshot A")

    changed = [d for d in result.element_deltas if d.status == "changed"]
    if changed:
        top = sorted(changed, key=lambda d: abs(d.budget_delta), reverse=True)[:3]
        for d in top:
            result.insights.append(
                f"{d.cbs_code}: ${d.budget_delta / 1e6:+.1f}M ({d.variance_pct:+.1f}%)"
            )

    if abs(result.budget_variance_pct) > 10:
        result.insights.append(
            "Total variance exceeds 10% — significant scope or cost shift; "
            "review per AACE RP 29R-03 §5.3."
        )
