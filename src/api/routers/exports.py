# MIT License
# Copyright (c) 2026 Vitor Maia Rodovalho
"""Exports router — XER, Excel, JSON, CSV export + activity search."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import asdict
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import PlainTextResponse, StreamingResponse

from ..auth import optional_auth
from ..deps import get_store

from src.analytics.float_trends import compute_float_entropy
from src.parser.models import ParsedSchedule

router = APIRouter()


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------


def _serialize_for_json(obj: Any) -> Any:
    """Recursively convert dataclass / datetime objects to JSON-safe dicts."""
    from datetime import date, datetime

    if obj is None or isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _serialize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_serialize_for_json(i) for i in obj]
    if hasattr(obj, "__dataclass_fields__"):
        return {k: _serialize_for_json(v) for k, v in asdict(obj).items()}
    if hasattr(obj, "model_dump"):
        return obj.model_dump()
    return str(obj)


def _build_export_data(
    schedule: ParsedSchedule,
    user_id: str | None,
) -> dict[str, Any]:
    """Build a comprehensive data bundle for JSON/CSV export.

    Runs DCMA, health score, and float analysis automatically.
    Returns a dict with all analysis results.
    """
    from src.analytics.dcma14 import DCMA14Analyzer
    from src.analytics.health_score import HealthScoreCalculator

    project_name = ""
    data_date = None
    if schedule.projects:
        proj = schedule.projects[0]
        project_name = proj.proj_short_name or ""
        data_date = proj.last_recalc_date or proj.sum_data_date

    dcma = DCMA14Analyzer(schedule)
    dcma_result = dcma.analyze()

    calc = HealthScoreCalculator(schedule)
    health = calc.calculate()

    # Activity summary
    activities_data = []
    for a in schedule.activities:
        activities_data.append(
            {
                "task_id": a.task_id,
                "task_code": a.task_code,
                "task_name": a.task_name,
                "task_type": a.task_type,
                "status_code": a.status_code,
                "total_float_hr": a.total_float_hr_cnt,
                "remaining_duration_hr": a.remain_drtn_hr_cnt,
                "original_duration_hr": a.target_drtn_hr_cnt,
                "physical_pct_complete": getattr(a, "phys_complete_pct", None),
                "actual_start": _serialize_for_json(a.act_start_date),
                "actual_finish": _serialize_for_json(a.act_end_date),
                "early_start": _serialize_for_json(a.early_start_date),
                "early_finish": _serialize_for_json(a.early_end_date),
                "late_start": _serialize_for_json(getattr(a, "late_start_date", None)),
                "late_finish": _serialize_for_json(getattr(a, "late_end_date", None)),
                "wbs_id": getattr(a, "wbs_id", None),
            }
        )

    # DCMA metrics
    dcma_metrics = []
    if hasattr(dcma_result, "metrics"):
        for m in dcma_result.metrics:
            dcma_metrics.append(
                {
                    "number": m.number,
                    "name": m.name,
                    "value": m.value,
                    "threshold": m.threshold,
                    "unit": m.unit,
                    "passed": m.passed,
                    "description": getattr(m, "description", ""),
                }
            )

    entropy = compute_float_entropy(schedule)

    return {
        "export_version": "1.0",
        "generator": "MeridianIQ",
        "project": {
            "name": project_name,
            "data_date": _serialize_for_json(data_date),
            "activity_count": len(schedule.activities),
            "relationship_count": len(schedule.relationships),
            "wbs_count": len(schedule.wbs_nodes),
            "calendar_count": len(schedule.calendars),
        },
        "health_score": {
            "overall": health.overall,
            "rating": health.rating,
            "trend": health.trend_arrow,
            "dcma_raw": health.dcma_raw,
            "float_raw": health.float_raw,
            "logic_raw": health.logic_raw,
            "trend_raw": health.trend_raw,
            "dcma_component": health.dcma_component,
            "float_component": health.float_component,
            "logic_component": health.logic_component,
            "trend_component": health.trend_component,
        },
        "dcma": {
            "overall_score": dcma_result.overall_score
            if hasattr(dcma_result, "overall_score")
            else 0,
            "passed_count": dcma_result.passed_count if hasattr(dcma_result, "passed_count") else 0,
            "failed_count": dcma_result.failed_count if hasattr(dcma_result, "failed_count") else 0,
            "metrics": dcma_metrics,
        },
        "float_entropy": {
            "entropy": entropy.entropy,
            "normalised_entropy": entropy.normalised_entropy,
            "max_entropy": entropy.max_entropy,
            "bucket_count": entropy.bucket_count,
            "distribution": entropy.distribution,
            "interpretation": entropy.interpretation,
        },
        "activities": activities_data,
        "relationships": [
            {
                "task_id": r.task_id,
                "pred_task_id": r.pred_task_id,
                "pred_type": r.pred_type,
                "lag_hr": r.lag_hr_cnt,
            }
            for r in schedule.relationships
        ],
    }


# ------------------------------------------------------------------
# XER Export
# ------------------------------------------------------------------


@router.get("/api/v1/projects/{project_id}/export/xer")
def export_xer(
    project_id: str,
    _user: object = Depends(optional_auth),
) -> dict:
    """Export a project schedule to XER format for P6 import.

    Returns the XER file content as a downloadable string.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    from src.export.xer_writer import XERWriter

    content = XERWriter(schedule).write()

    return PlainTextResponse(
        content=content,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={project_id}.xer"},
    )


# ------------------------------------------------------------------
# Excel Export
# ------------------------------------------------------------------


@router.get("/api/v1/projects/{project_id}/export/excel")
def export_excel(
    project_id: str,
    _user: object = Depends(optional_auth),
):
    """Export project schedule data as an Excel workbook.

    Includes sheets for: Activities, Relationships, WBS, and Summary.
    PM and Owner personas use this for custom reporting in Excel.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found or openpyxl is unavailable.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(
            status_code=501, detail="Excel export requires openpyxl (pip install openpyxl)"
        )

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["MeridianIQ Schedule Export"])
    ws_summary.append(
        ["Project", schedule.projects[0].proj_short_name if schedule.projects else ""]
    )
    ws_summary.append(["Activities", len(schedule.activities)])
    ws_summary.append(["Relationships", len(schedule.relationships)])
    ws_summary.append(["WBS Elements", len(schedule.wbs_nodes)])
    ws_summary.append(["Calendars", len(schedule.calendars)])
    ws_summary.append(["Parser Version", schedule.parser_version])

    # Activities sheet
    ws_act = wb.create_sheet("Activities")
    ws_act.append(
        [
            "Task ID",
            "Task Code",
            "Task Name",
            "Type",
            "Status",
            "Total Float (hrs)",
            "Remaining Duration (hrs)",
            "Original Duration (hrs)",
            "Actual Start",
            "Actual Finish",
            "Early Start",
            "Early Finish",
        ]
    )
    for t in schedule.activities:
        ws_act.append(
            [
                t.task_id,
                t.task_code,
                t.task_name,
                t.task_type,
                t.status_code,
                t.total_float_hr_cnt,
                t.remain_drtn_hr_cnt,
                t.target_drtn_hr_cnt,
                str(t.act_start_date) if t.act_start_date else "",
                str(t.act_end_date) if t.act_end_date else "",
                str(t.early_start_date) if t.early_start_date else "",
                str(t.early_end_date) if t.early_end_date else "",
            ]
        )

    # Relationships sheet
    ws_rel = wb.create_sheet("Relationships")
    ws_rel.append(["Task ID", "Predecessor ID", "Type", "Lag (hrs)"])
    for r in schedule.relationships:
        ws_rel.append([r.task_id, r.pred_task_id, r.pred_type, r.lag_hr_cnt])

    # WBS sheet
    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.append(["WBS ID", "Parent WBS ID", "Short Name", "Name"])
    for w in schedule.wbs_nodes:
        ws_wbs.append([w.wbs_id, w.parent_wbs_id, w.wbs_short_name, w.wbs_name])

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    proj_name = schedule.projects[0].proj_short_name if schedule.projects else project_id
    filename = f"meridianiq-{proj_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ------------------------------------------------------------------
# AIA G703 Continuation Sheet Export
# ------------------------------------------------------------------


@router.get("/api/v1/projects/{project_id}/export/aia-g703")
def export_aia_g703(
    project_id: str,
    snapshot_id: str,
    retainage_pct: float = 0.10,
    application_number: int = 1,
    period_to: str = "",
    _user: object = Depends(optional_auth),
):
    """Export CBS snapshot as an AIA G703 Continuation Sheet workbook.

    Produces the schedule of values attached to an AIA G702 Application
    for Payment. CBS elements map to line items; retainage, application
    number, and period-to-date are caller inputs (not stored with the
    CBS upload).

    Args:
        project_id: The stored project identifier.
        snapshot_id: CBS snapshot id from ``/cost/snapshots``.
        retainage_pct: Retainage fraction (default 0.10 = 10 %).
        application_number: Billing application sequence (default 1).
        period_to: Period-end date string (default empty).

    Raises:
        HTTPException: 404 if project or snapshot not retrievable;
        501 if openpyxl is not available.

    Reference: AIA Document G703™ — Continuation Sheet.
    """
    store = get_store()
    user_id = _user["id"] if _user else None

    schedule = store.get(project_id, user_id=user_id) if hasattr(store, "get") else None
    if schedule is None and hasattr(store, "get_parsed_schedule"):
        schedule = store.get_parsed_schedule(project_id)

    project_name = (
        schedule.projects[0].proj_short_name if schedule and schedule.projects else project_id
    )

    snapshot = None
    if hasattr(store, "get_cost_snapshot"):
        snapshot = store.get_cost_snapshot(project_id, snapshot_id, user_id=user_id)
    if snapshot is None:
        raise HTTPException(status_code=404, detail=f"CBS snapshot not retrievable: {snapshot_id}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=501, detail="Excel export requires openpyxl (pip install openpyxl)"
        )

    from src.analytics.aia_g703 import build_g703_from_cbs

    sheet = build_g703_from_cbs(
        snapshot,
        project_name=project_name,
        application_number=application_number,
        period_to=period_to,
        retainage_pct=retainage_pct,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "G703 Continuation"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    # Header block (AIA G703 top-of-form fields)
    ws["A1"] = "AIA Document G703™ — Continuation Sheet"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A1"].alignment = center

    ws["A3"] = "Project:"
    ws["B3"] = sheet.project_name
    ws["A3"].font = bold
    ws["F3"] = "Application No:"
    ws["G3"] = sheet.application_number
    ws["F3"].font = bold
    ws["A4"] = "Period To:"
    ws["B4"] = sheet.period_to
    ws["A4"].font = bold
    ws["F4"] = "Retainage %:"
    ws["G4"] = f"{sheet.retainage_pct * 100:.1f}%"
    ws["F4"].font = bold

    # Column headers (row 6)
    headers = [
        "A\nItem No.",
        "B\nDescription of Work",
        "C\nScheduled Value",
        "D\nFrom Previous\nApplication",
        "E\nThis Period",
        "F\nMaterials\nPresently Stored",
        "G = D+E+F\nTotal Completed\n& Stored",
        "H = G/C\n% Complete",
        "I = C-G\nBalance to Finish",
        "J\nRetainage",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    ws.row_dimensions[6].height = 48
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 40
    for col_letter in ("C", "D", "E", "F", "G", "I", "J"):
        ws.column_dimensions[col_letter].width = 16
    ws.column_dimensions["H"].width = 12

    # Line items
    row = 7
    for li in sheet.line_items:
        ws.cell(row=row, column=1, value=li.line_number)
        ws.cell(row=row, column=2, value=li.description)
        ws.cell(row=row, column=3, value=li.scheduled_value).alignment = right
        ws.cell(row=row, column=4, value=li.previous_application_value).alignment = right
        ws.cell(row=row, column=5, value=li.this_period_value).alignment = right
        ws.cell(row=row, column=6, value=li.materials_presently_stored).alignment = right
        ws.cell(row=row, column=7, value=li.total_completed_and_stored).alignment = right
        ws.cell(row=row, column=8, value=f"{li.percent_complete:.1f}%").alignment = right
        ws.cell(row=row, column=9, value=li.balance_to_finish).alignment = right
        ws.cell(row=row, column=10, value=li.retainage).alignment = right
        row += 1

    # Totals
    ws.cell(row=row, column=1, value="").font = bold
    ws.cell(row=row, column=2, value="GRAND TOTAL").font = bold
    totals = [
        sheet.total_scheduled_value,
        sheet.total_previous_application,
        sheet.total_this_period,
        sheet.total_materials_stored,
        sheet.total_completed_and_stored,
        None,  # percent complete at grand-total level
        sheet.total_balance_to_finish,
        sheet.total_retainage,
    ]
    pct = (
        (sheet.total_completed_and_stored / sheet.total_scheduled_value * 100)
        if sheet.total_scheduled_value > 0
        else 0.0
    )
    totals[5] = f"{pct:.1f}%"
    for offset, value in enumerate(totals, start=3):
        cell = ws.cell(row=row, column=offset, value=value)
        cell.font = bold
        cell.alignment = right
        cell.fill = header_fill

    # Methodology note
    ws.cell(
        row=row + 2,
        column=1,
        value=(
            "Prepared per AIA G703 Continuation Sheet format. Retainage %, "
            "application #, and period-to-date supplied by requester; "
            "percent complete derived from CBS snapshot progress."
        ),
    ).font = Font(italic=True, size=9)
    ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"meridianiq-aia-g703-{project_name}-app{application_number:03d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ------------------------------------------------------------------
# JSON Export
# ------------------------------------------------------------------


@router.get("/api/v1/projects/{project_id}/export/json")
def export_json(
    project_id: str,
    _user: object = Depends(optional_auth),
) -> StreamingResponse:
    """Export project schedule data and analysis results as JSON.

    Includes project metadata, DCMA 14-point assessment, health score,
    activities with dates/float, and relationships. Designed for
    researchers and external analysis tools.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    data = _build_export_data(schedule, user_id)

    json_bytes = json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")

    proj_name = data["project"]["name"] or project_id
    filename = f"meridianiq-{proj_name}.json"

    return StreamingResponse(
        io.BytesIO(json_bytes),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ------------------------------------------------------------------
# CSV Export
# ------------------------------------------------------------------


@router.get("/api/v1/projects/{project_id}/export/csv")
def export_csv(
    project_id: str,
    dataset: str = "activities",
    _user: object = Depends(optional_auth),
) -> StreamingResponse:
    """Export project data as CSV.

    Supports multiple datasets via the ``dataset`` query parameter:
    - ``activities`` (default): All activities with dates, float, status
    - ``dcma``: DCMA 14-point check results
    - ``relationships``: All predecessor relationships

    Args:
        project_id: The stored project identifier.
        dataset: Which dataset to export (activities, dcma, relationships).

    Raises:
        HTTPException: If the project is not found or dataset is invalid.
    """
    valid_datasets = {"activities", "dcma", "relationships"}
    if dataset not in valid_datasets:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid dataset '{dataset}'. Valid: {', '.join(sorted(valid_datasets))}",
        )

    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    data = _build_export_data(schedule, user_id)
    buf = io.StringIO()
    writer = csv.writer(buf)

    if dataset == "activities":
        headers = [
            "task_id",
            "task_code",
            "task_name",
            "task_type",
            "status_code",
            "total_float_hr",
            "remaining_duration_hr",
            "original_duration_hr",
            "physical_pct_complete",
            "actual_start",
            "actual_finish",
            "early_start",
            "early_finish",
            "late_start",
            "late_finish",
            "wbs_id",
        ]
        writer.writerow(headers)
        for a in data["activities"]:
            writer.writerow([a.get(h, "") for h in headers])

    elif dataset == "dcma":
        headers = ["number", "name", "value", "threshold", "unit", "passed", "description"]
        writer.writerow(headers)
        for m in data["dcma"]["metrics"]:
            writer.writerow([m.get(h, "") for h in headers])

    elif dataset == "relationships":
        headers = ["task_id", "pred_task_id", "pred_type", "lag_hr"]
        writer.writerow(headers)
        for r in data["relationships"]:
            writer.writerow([r.get(h, "") for h in headers])

    proj_name = data["project"]["name"] or project_id
    filename = f"meridianiq-{proj_name}-{dataset}.csv"

    csv_bytes = buf.getvalue().encode("utf-8")

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ------------------------------------------------------------------
# Activity Search (for TIA autocomplete)
# ------------------------------------------------------------------


@router.get("/api/v1/projects/{project_id}/activities")
def search_activities(
    project_id: str,
    q: str = "",
    limit: int = 20,
    _user: object = Depends(optional_auth),
) -> dict:
    """Search activities by ID or name.  Used by the TIA activity picker.

    Args:
        project_id: The stored project identifier.
        q: Optional search query matched against task_code and task_name.
        limit: Maximum number of results to return (default 20).

    Returns:
        Dict with ``activities`` list and ``total`` count of all activities.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        # Fallback: try get_parsed_schedule (SupabaseStore path)
        if hasattr(store, "get_parsed_schedule"):
            schedule = store.get_parsed_schedule(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    total = len(schedule.activities)
    activities: list[dict] = []
    q_lower = q.strip().lower()

    for task in schedule.activities:
        entry = {
            "task_code": task.task_code,
            "task_name": task.task_name,
            "task_type": task.task_type,
            "wbs_id": task.wbs_id,
            "status_code": task.status_code,
        }
        if q_lower:
            searchable = f"{task.task_code} {task.task_name}".lower()
            if q_lower not in searchable:
                continue
        activities.append(entry)
        if len(activities) >= limit:
            break

    return {"activities": activities, "total": total}
