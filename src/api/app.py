# MIT License
# Copyright (c) 2026 Vitor Maia Rodovalho
"""FastAPI application for MeridianIQ.

Provides REST endpoints for uploading XER files, querying parsed
schedule data, running DCMA assessments, CPM analysis, comparing
schedule versions, and forensic (CPA/window) analysis.
"""

from __future__ import annotations

import csv
import io
import json
import os
import tempfile
from dataclasses import asdict
from pathlib import Path
from typing import Any

import sentry_sdk

if dsn := os.environ.get("SENTRY_DSN"):
    sentry_sdk.init(
        dsn=dsn,
        traces_sample_rate=0.1,
        environment=os.environ.get("ENVIRONMENT", "development"),
        release="meridianiq-api@1.0.0-dev",
    )

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse

from src.analytics.comparison import ScheduleComparison
from src.analytics.report_generator import ReportGenerator
from src.analytics.contract import ContractComplianceChecker
from src.analytics.cpm import CPMCalculator
from src.analytics.dcma14 import DCMA14Analyzer
from src.analytics.early_warning import EarlyWarningEngine
from src.analytics.evm import EVMAnalyzer
from src.analytics.float_trends import (
    FloatTrendAnalyzer,
    compute_float_entropy,
    compute_constraint_accumulation,
)
from src.analytics.health_score import HealthScoreCalculator
from src.analytics.forensics import ForensicAnalyzer
from src.analytics.risk import (
    DistributionType,
    DurationRisk,
    MonteCarloSimulator,
    RiskEvent,
    SimulationConfig,
)
from src.analytics.tia import (
    DelayFragment,
    FragmentActivity,
    ResponsibleParty,
    TimeImpactAnalyzer,
)
from src.parser.models import ParsedSchedule
from src.parser.xer_reader import XERReader

from .schemas import (
    ActivityChangeSchema,
    ActivitySchema,
    ActivityStatusSummary,
    CodeRestructuringSchema,
    CompareRequest,
    CompareResponse,
    ComplianceCheckSchema,
    ContractCheckRequest,
    ContractCheckResponse,
    ContractProvisionSchema,
    ContractProvisionsResponse,
    CreateTimelineRequest,
    CriticalPathActivity,
    CriticalPathResponse,
    DelayFragmentSchema,
    DelayTrendPoint,
    DelayTrendResponse,
    EVMAnalysisSchema,
    EVMAnalysisSummarySchema,
    EVMListResponse,
    EVMMetricsSchema,
    ForecastResponse,
    FloatBucket,
    FloatChangeSchema,
    FloatDistributionResponse,
    FragmentActivitySchema,
    HealthClassificationSchema,
    HealthResponse,
    ManipulationFlagSchema,
    MatchStatsSchema,
    MetricSchema,
    MilestoneSchema,
    MilestonesResponse,
    ProjectDetailResponse,
    ProjectListItem,
    ProjectListResponse,
    ProjectSummary,
    RelationshipChangeSchema,
    RelationshipSchema,
    RelationshipTypeSummary,
    SCurvePointSchema,
    SCurveResponse,
    TIAAnalysisSchema,
    TIAAnalysisSummarySchema,
    TIAAnalyzeRequest,
    TIAListResponse,
    TIAResultSchema,
    TIASummaryResponse,
    TimelineDetailSchema,
    TimelineListResponse,
    TimelineSummarySchema,
    ValidationResponse,
    WBSDrillResponse,
    WBSLevelCount,
    WBSMetricsSchema,
    WBSStats,
    WindowSchema,
    # Risk schemas
    CriticalityEntrySchema,
    CriticalityResponse,
    HistogramBinSchema,
    HistogramResponse,
    PValueSchema,
    RiskSCurvePointSchema,
    RiskSCurveResponse,
    RunSimulationRequest,
    SensitivityEntrySchema,
    SimulationListResponse,
    SimulationResultSchema,
    SimulationSummarySchema,
    TornadoResponse,
    # Intelligence v0.8 schemas
    ActivityFloatTrendSchema,
    AlertSchema,
    AlertsResponse,
    DashboardKPIs,
    FloatTrendResponse,
    GenerateReportRequest,
    GenerateReportResponse,
    ScheduleHealthResponse,
)
from .auth import optional_auth, require_auth
from .storage import EVMStore, ProjectStore, ReportStore, RiskStore, TIAStore, TimelineStore
from src.database.store import get_store as _get_db_store

app = FastAPI(
    title="MeridianIQ",
    description="The intelligence standard for project schedules",
    version="1.3.0-dev",
)

# Rate limiting
try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.errors import RateLimitExceeded
    from slowapi.util import get_remote_address

    limiter = Limiter(key_func=get_remote_address, default_limits=["60/minute"])
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
    _has_rate_limiter = True
except ImportError:
    _has_rate_limiter = False

# CORS — whitelist known origins (not wildcard)
_CORS_ORIGINS = [
    "http://localhost:5173",
    "http://localhost:4321",
    "https://meridianiq.vitormr.dev",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
    max_age=3600,
)


# Security headers middleware
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    """Add standard security headers to all responses."""
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    if request.url.scheme == "https":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


from .organizations import router as org_router  # noqa: E402

app.include_router(org_router)


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception) -> JSONResponse:
    """Catch-all exception handler that ensures CORS headers on error responses."""
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc)},
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "*",
            "Access-Control-Allow-Headers": "*",
        },
    )


# Global stores (singletons for the app lifetime)
# _store uses the database factory: SupabaseStore in production, InMemoryStore in dev
_store = _get_db_store()
_timeline_store = TimelineStore()
_tia_store = TIAStore()
_evm_store = EVMStore()
_risk_store = RiskStore()
_report_store = ReportStore()


def get_store():
    """Return the global project store.

    Exposed as a function so tests can replace it via monkeypatching.
    """
    return _store


def get_timeline_store() -> TimelineStore:
    """Return the global timeline store.

    Exposed as a function so tests can replace it via monkeypatching.
    """
    return _timeline_store


def get_tia_store() -> TIAStore:
    """Return the global TIA store.

    Exposed as a function so tests can replace it via monkeypatching.
    """
    return _tia_store


def get_evm_store() -> EVMStore:
    """Return the global EVM store.

    Exposed as a function so tests can replace it via monkeypatching.
    """
    return _evm_store


def get_risk_store() -> RiskStore:
    """Return the global risk simulation store.

    Exposed as a function so tests can replace it via monkeypatching.
    """
    return _risk_store


def get_report_store() -> ReportStore:
    """Return the global report store.

    Exposed as a function so tests can replace it via monkeypatching.
    """
    return _report_store


# ------------------------------------------------------------------
# Health
# ------------------------------------------------------------------


@app.get("/health")
async def root_health():
    return {"status": "ok", "version": "1.0.0-dev"}


@app.get("/api/v1/health", response_model=HealthResponse)
def health_check() -> HealthResponse:
    """Health check endpoint."""
    return HealthResponse()


# ------------------------------------------------------------------
# Demo (unauthenticated sample data)
# ------------------------------------------------------------------


@app.get("/api/v1/demo/project")
def demo_project():
    """Return a pre-analyzed demo project from the sample XER fixture.

    No authentication required. Used by the landing page "Try with sample data" flow.
    """
    sample_path = Path(__file__).resolve().parents[2] / "tests" / "fixtures" / "sample.xer"
    if not sample_path.exists():
        raise HTTPException(status_code=404, detail="Demo data not available")

    from src.parser.xer_reader import XerReader
    from src.analytics.cpm import CPMEngine

    reader = XerReader()
    schedule = reader.parse(sample_path.read_text(encoding="utf-8"))
    analyzer = DCMA14Analyzer(schedule)
    dcma = analyzer.analyze()

    cpm = CPMEngine(schedule)
    cpm_result = cpm.calculate()

    return {
        "project": {
            "name": schedule.project.proj_short_name if schedule.project else "Demo Project",
            "activity_count": len(schedule.activities),
            "relationship_count": len(schedule.relationships),
            "calendar_count": len(schedule.calendars),
            "wbs_count": len(schedule.wbs_nodes),
        },
        "validation": {
            "overall_score": dcma.overall_score,
            "passed_count": dcma.passed_count,
            "failed_count": dcma.failed_count,
            "metrics": [
                {
                    "number": m.number,
                    "name": m.name,
                    "value": round(m.value, 1),
                    "threshold": m.threshold,
                    "unit": m.unit,
                    "passed": m.passed,
                    "direction": m.direction,
                }
                for m in dcma.metrics
            ],
        },
        "critical_path": {
            "length": len(cpm_result.critical_path),
            "activities": [
                {
                    "task_code": a.task_code,
                    "task_name": a.task_name,
                    "total_float": round((a.total_float_hr_cnt or 0) / 8, 1),
                }
                for a in cpm_result.critical_path[:20]
            ],
        },
    }


# ------------------------------------------------------------------
# Upload
# ------------------------------------------------------------------

# Sandbox project tracking
# In-memory for dev; production should use is_sandbox DB column (migration 010).
# TODO(v2.0): migrate to SECURITY DEFINER RPC that sets is_sandbox on schedule_uploads.
_sandbox_projects: set[str] = set()


@app.post("/api/v1/upload", response_model=ProjectSummary)
async def upload_xer(
    file: UploadFile = File(...),
    is_sandbox: bool = Form(False),
    _user: object = Depends(optional_auth),
) -> ProjectSummary:
    """Upload a schedule file (XER or MS Project XML), parse it, and store the result.

    Supports:
    - Primavera P6 XER files (.xer)
    - Microsoft Project XML files (.xml)

    Returns:
        A summary of the parsed project.

    Raises:
        HTTPException: If the file format is not supported.
    """
    filename = (file.filename or "").lower()
    is_xer = filename.endswith(".xer")
    is_xml = filename.endswith(".xml")

    if not is_xer and not is_xml:
        raise HTTPException(
            status_code=400,
            detail="Unsupported format. Upload a .xer (Primavera P6) or .xml (Microsoft Project) file.",
        )

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty file")

    # Enforce upload size limit (50 MB)
    MAX_UPLOAD_SIZE = 50 * 1024 * 1024
    if len(file_bytes) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large ({len(file_bytes) / 1024 / 1024:.1f} MB). "
            f"Maximum allowed: {MAX_UPLOAD_SIZE / 1024 / 1024:.0f} MB.",
        )

    # Parse based on format
    try:
        if is_xml:
            from src.parser.msp_reader import MSPReader

            msp_reader = MSPReader()
            schedule = msp_reader.parse(file_bytes.decode("utf-8"))
            xer_bytes = file_bytes  # Store XML as-is
        else:
            xer_bytes = file_bytes
            with tempfile.NamedTemporaryFile(suffix=".xer", delete=False) as tmp:
                tmp.write(xer_bytes)
                tmp_path = Path(tmp.name)

            reader = XERReader(tmp_path)
            schedule = reader.parse()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse XER file: {exc}")
    finally:
        try:
            tmp_path.unlink()
        except Exception:
            pass

    if not schedule.activities:
        raise HTTPException(status_code=400, detail="XER file contains no activities")

    store = get_store()
    user_id = _user["id"] if _user else None
    project_id = store.add(schedule, xer_bytes, user_id=user_id)

    # Track sandbox status
    if is_sandbox:
        _sandbox_projects.add(project_id)

    data_date = None
    name = ""
    if schedule.projects:
        proj = schedule.projects[0]
        name = proj.proj_short_name
        dd = proj.last_recalc_date or proj.sum_data_date
        if dd:
            data_date = dd.isoformat()

    return ProjectSummary(
        project_id=project_id,
        name=name,
        activity_count=len(schedule.activities),
        relationship_count=len(schedule.relationships),
        calendar_count=len(schedule.calendars),
        wbs_count=len(schedule.wbs_nodes),
        data_date=data_date,
    )


# ------------------------------------------------------------------
# Project listing
# ------------------------------------------------------------------


@app.get("/api/v1/projects", response_model=ProjectListResponse)
def list_projects(
    include_sandbox: bool = False,
    _user: object = Depends(optional_auth),
) -> ProjectListResponse:
    """List all uploaded projects.

    By default, sandbox projects are hidden. Pass ``include_sandbox=true``
    to include them (only visible to the owner regardless).
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    all_projects = store.list_all(user_id=user_id)
    if not include_sandbox:
        all_projects = [p for p in all_projects if p["project_id"] not in _sandbox_projects]
    items = [ProjectListItem(**p) for p in all_projects]
    return ProjectListResponse(projects=items)


@app.put("/api/v1/projects/{project_id}/sandbox")
def toggle_sandbox(
    project_id: str,
    _user: object = Depends(optional_auth),
) -> dict:
    """Toggle sandbox mode for a project.

    Sandbox projects are hidden from project listings and org views.
    Only the owner can see and toggle sandbox status.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    if project_id in _sandbox_projects:
        _sandbox_projects.discard(project_id)
        is_sandbox = False
    else:
        _sandbox_projects.add(project_id)
        is_sandbox = True

    return {"project_id": project_id, "is_sandbox": is_sandbox}


# ------------------------------------------------------------------
# Programs (group uploads under programs with revision tracking)
# ------------------------------------------------------------------


@app.get("/api/v1/programs")
def list_programs(_user: object = Depends(optional_auth)):
    """Return all programs with latest revision info."""
    store = get_store()
    user_id = _user["id"] if _user else None
    programs = store.get_programs(user_id=user_id)
    return {"programs": programs}


@app.get("/api/v1/programs/{program_id}")
def get_program_detail(program_id: str, _user: object = Depends(optional_auth)):
    """Return a program with all its revisions."""
    store = get_store()
    user_id = _user["id"] if _user else None
    revisions = store.get_program_revisions(program_id, user_id=user_id)
    # Also get the program metadata
    programs = store.get_programs(user_id=user_id)
    program = None
    for p in programs:
        if p["id"] == program_id:
            program = p
            break
    if program is None:
        raise HTTPException(status_code=404, detail="Program not found")
    return {"program": program, "revisions": revisions}


@app.put("/api/v1/programs/{program_id}")
def update_program(program_id: str, body: dict, _user: object = Depends(optional_auth)):
    """Rename or update a program."""
    store = get_store()
    user_id = _user["id"] if _user else None
    updated = store.update_program(program_id, body, user_id=user_id)
    if updated is None:
        raise HTTPException(status_code=404, detail="Program not found")
    return {"program": updated}


@app.get("/api/v1/programs/{program_id}/trends")
def get_program_trends(program_id: str, _user: object = Depends(optional_auth)):
    """Trend data across all revisions for charting."""
    store = get_store()
    user_id = _user["id"] if _user else None
    revisions = (
        store.get_program_revisions(program_id, user_id)
        if hasattr(store, "get_program_revisions")
        else []
    )
    if not revisions:
        raise HTTPException(status_code=404, detail="Program not found or no revisions")

    # Sort ascending by revision_number
    revisions.sort(key=lambda r: r.get("revision_number", 0))

    trends: dict = {
        "revision_count": len(revisions),
        "labels": [],
        "health_scores": [],
        "dcma_scores": [],
        "alert_counts": [],
        "activity_counts": [],
        "revisions": [],
    }

    for rev in revisions:
        results = rev.get("analysis_results") or {}
        health = results.get("health", {})
        dcma = results.get("dcma", {})
        alerts = results.get("alerts", {})

        label = rev.get("data_date") or f"Rev {rev.get('revision_number', '?')}"
        trends["labels"].append(str(label))
        trends["health_scores"].append(health.get("score") if health else None)
        dcma_score = dcma.get("score") if dcma else None
        if dcma_score is None and dcma:
            dcma_score = dcma.get("pass_rate")
        trends["dcma_scores"].append(dcma_score)
        if alerts and isinstance(alerts.get("alerts"), list):
            trends["alert_counts"].append(len(alerts["alerts"]))
        elif alerts:
            trends["alert_counts"].append(alerts.get("count", 0))
        else:
            trends["alert_counts"].append(None)
        trends["activity_counts"].append(rev.get("activity_count"))
        trends["revisions"].append(
            {
                "id": rev.get("id"),
                "revision_number": rev.get("revision_number"),
                "data_date": rev.get("data_date"),
                "filename": rev.get("filename"),
            }
        )

    return trends


# ------------------------------------------------------------------
# Project detail
# ------------------------------------------------------------------


@app.get("/api/v1/projects/{project_id}", response_model=ProjectDetailResponse)
def get_project(project_id: str, _user: object = Depends(optional_auth)) -> ProjectDetailResponse:
    """Get full project data for a given project_id.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    data_date = None
    name = ""
    if schedule.projects:
        proj = schedule.projects[0]
        name = proj.proj_short_name
        dd = proj.last_recalc_date or proj.sum_data_date
        if dd:
            data_date = dd.isoformat()

    activities = [
        ActivitySchema(
            task_id=t.task_id,
            task_code=t.task_code,
            task_name=t.task_name,
            task_type=t.task_type,
            status_code=t.status_code,
            total_float_hr_cnt=t.total_float_hr_cnt,
            remain_drtn_hr_cnt=t.remain_drtn_hr_cnt,
            target_drtn_hr_cnt=t.target_drtn_hr_cnt,
            act_start_date=t.act_start_date,
            act_end_date=t.act_end_date,
            early_start_date=t.early_start_date,
            early_end_date=t.early_end_date,
            late_start_date=t.late_start_date,
            late_end_date=t.late_end_date,
        )
        for t in schedule.activities
    ]

    relationships = [
        RelationshipSchema(
            task_id=r.task_id,
            pred_task_id=r.pred_task_id,
            pred_type=r.pred_type,
            lag_hr_cnt=r.lag_hr_cnt,
        )
        for r in schedule.relationships
    ]

    # Compute WBS hierarchy statistics
    wbs_stats = _compute_wbs_stats(schedule)

    # Compute summary counts server-side (avoids client iterating full arrays)
    complete = sum(1 for t in schedule.activities if t.status_code.upper() == "TK_COMPLETE")
    active = sum(1 for t in schedule.activities if t.status_code.upper() == "TK_ACTIVE")
    activity_summary = ActivityStatusSummary(
        total=len(schedule.activities),
        complete=complete,
        in_progress=active,
        not_started=len(schedule.activities) - complete - active,
    )

    fs = ff = ss = sf = 0
    for r in schedule.relationships:
        pt = r.pred_type.upper() if r.pred_type else ""
        if pt == "PR_FS":
            fs += 1
        elif pt == "PR_FF":
            ff += 1
        elif pt == "PR_SS":
            ss += 1
        elif pt == "PR_SF":
            sf += 1
    relationship_summary = RelationshipTypeSummary(
        total=len(schedule.relationships),
        fs=fs,
        ff=ff,
        ss=ss,
        sf=sf,
    )

    return ProjectDetailResponse(
        project_id=project_id,
        name=name,
        data_date=data_date,
        activities=activities,
        relationships=relationships,
        wbs_stats=wbs_stats,
        activity_summary=activity_summary,
        relationship_summary=relationship_summary,
    )


def _compute_wbs_stats(schedule: ParsedSchedule) -> WBSStats:
    """Compute WBS hierarchy depth and activity distribution."""
    wbs_nodes = schedule.wbs_nodes
    if not wbs_nodes:
        return WBSStats()

    # Build parent→children map and find root(s)
    child_ids: set[str] = set()
    for w in wbs_nodes:
        if w.parent_wbs_id and w.parent_wbs_id != w.wbs_id:
            child_ids.add(w.wbs_id)

    # Calculate depth for each WBS node via BFS from roots
    roots = [w for w in wbs_nodes if w.wbs_id not in child_ids or w.proj_node_flag.upper() == "Y"]
    levels: dict[str, int] = {}
    queue = [(r.wbs_id, 1) for r in roots]
    for wbs_id, level in queue:
        if wbs_id in levels:
            continue
        levels[wbs_id] = level
        # Find children
        for w in wbs_nodes:
            if w.parent_wbs_id == wbs_id and w.wbs_id != wbs_id:
                queue.append((w.wbs_id, level + 1))

    # Assign level 1 to any unvisited nodes (orphans)
    for w in wbs_nodes:
        if w.wbs_id not in levels:
            levels[w.wbs_id] = 1

    max_depth = max(levels.values()) if levels else 0

    # Count by level
    level_counts: dict[int, int] = {}
    for lvl in levels.values():
        level_counts[lvl] = level_counts.get(lvl, 0) + 1
    by_level = [WBSLevelCount(level=lvl, count=cnt) for lvl, cnt in sorted(level_counts.items())]

    # Activities per WBS node
    act_per_wbs: dict[str, int] = {w.wbs_id: 0 for w in wbs_nodes}
    for t in schedule.activities:
        if t.wbs_id in act_per_wbs:
            act_per_wbs[t.wbs_id] += 1

    counts = list(act_per_wbs.values())
    no_activities = sum(1 for c in counts if c == 0)
    non_zero = [c for c in counts if c > 0]

    return WBSStats(
        total_elements=len(wbs_nodes),
        max_depth=max_depth,
        by_level=by_level,
        avg_activities_per_wbs=round(sum(counts) / len(counts), 1) if counts else 0.0,
        min_activities_per_wbs=min(non_zero) if non_zero else 0,
        max_activities_per_wbs=max(non_zero) if non_zero else 0,
        wbs_with_no_activities=no_activities,
    )


# ------------------------------------------------------------------
# DCMA Validation
# ------------------------------------------------------------------


@app.get("/api/v1/projects/{project_id}/validation", response_model=ValidationResponse)
def get_validation(project_id: str, _user: object = Depends(optional_auth)) -> ValidationResponse:
    """Run DCMA 14-Point assessment for a project.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    analyzer = DCMA14Analyzer(schedule)
    dcma = analyzer.analyze()

    metrics = [
        MetricSchema(
            number=m.number,
            name=m.name,
            description=m.description,
            value=m.value,
            threshold=m.threshold,
            unit=m.unit,
            passed=m.passed,
            direction=m.direction,
            details=m.details,
        )
        for m in dcma.metrics
    ]

    return ValidationResponse(
        overall_score=dcma.overall_score,
        passed_count=dcma.passed_count,
        failed_count=dcma.failed_count,
        activity_count=dcma.activity_count,
        metrics=metrics,
    )


# ------------------------------------------------------------------
# Critical Path
# ------------------------------------------------------------------


@app.get(
    "/api/v1/projects/{project_id}/critical-path",
    response_model=CriticalPathResponse,
)
def get_critical_path(
    project_id: str, _user: object = Depends(optional_auth)
) -> CriticalPathResponse:
    """Compute and return the critical path for a project.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    cpm = CPMCalculator(schedule).calculate()

    cp_activities = [
        CriticalPathActivity(
            task_id=tid,
            task_code=cpm.activity_results[tid].task_code,
            task_name=cpm.activity_results[tid].task_name,
            duration=cpm.activity_results[tid].duration,
            early_start=cpm.activity_results[tid].early_start,
            early_finish=cpm.activity_results[tid].early_finish,
            total_float=cpm.activity_results[tid].total_float,
        )
        for tid in cpm.critical_path
        if tid in cpm.activity_results
    ]

    return CriticalPathResponse(
        project_duration=cpm.project_duration,
        critical_path=cp_activities,
        has_cycles=cpm.has_cycles,
    )


# ------------------------------------------------------------------
# Float Distribution
# ------------------------------------------------------------------


@app.get(
    "/api/v1/projects/{project_id}/float-distribution",
    response_model=FloatDistributionResponse,
)
def get_float_distribution(
    project_id: str, _user: object = Depends(optional_auth)
) -> FloatDistributionResponse:
    """Return float distribution buckets for a project.

    Buckets: critical (TF=0), near-critical (0-10d), moderate (10-20d),
    high (20-44d), excessive (>44d), negative.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Filter to countable activities (case-insensitive)
    countable = [
        t
        for t in schedule.activities
        if t.task_type.lower() not in ("tt_loe", "tt_wbs")
        and t.status_code.lower() != "tk_complete"
    ]

    total = len(countable)
    if total == 0:
        return FloatDistributionResponse(total_activities=0, buckets=[])

    # Define buckets in hours (8h/day)
    bucket_defs: list[tuple[str, float, float]] = [
        ("Negative (< 0)", float("-inf"), 0.0),
        ("Critical (= 0)", 0.0, 0.001),
        ("Near-Critical (0-10d)", 0.001, 80.001),
        ("Moderate (10-20d)", 80.001, 160.001),
        ("High (20-44d)", 160.001, 352.001),
        ("Excessive (> 44d)", 352.001, float("inf")),
    ]

    counts: dict[str, int] = {label: 0 for label, _, _ in bucket_defs}

    for task in countable:
        tf = task.total_float_hr_cnt
        if tf is None:
            continue
        for label, low, high in bucket_defs:
            if low <= tf < high:
                counts[label] += 1
                break

    buckets = [
        FloatBucket(
            range_label=label,
            count=counts[label],
            percentage=round((counts[label] / total) * 100, 1) if total else 0.0,
        )
        for label, _, _ in bucket_defs
    ]

    return FloatDistributionResponse(total_activities=total, buckets=buckets)


# ------------------------------------------------------------------
# Milestones
# ------------------------------------------------------------------


@app.get(
    "/api/v1/projects/{project_id}/milestones",
    response_model=MilestonesResponse,
)
def get_milestones(project_id: str, _user: object = Depends(optional_auth)) -> MilestonesResponse:
    """Return all milestone activities for a project.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Match milestones case-insensitively — P6 versions vary in case
    MILESTONE_TYPES = {"tt_mile", "tt_finmile"}
    milestones = [
        MilestoneSchema(
            task_id=t.task_id,
            task_code=t.task_code,
            task_name=t.task_name,
            task_type=t.task_type,
            status_code=t.status_code,
            act_start_date=t.act_start_date,
            act_end_date=t.act_end_date,
            early_start_date=t.early_start_date,
            early_end_date=t.early_end_date,
            target_start_date=t.target_start_date,
            target_end_date=t.target_end_date,
        )
        for t in schedule.activities
        if t.task_type.lower() in MILESTONE_TYPES
    ]

    return MilestonesResponse(milestones=milestones)


# ------------------------------------------------------------------
# Compare
# ------------------------------------------------------------------


@app.post("/api/v1/compare", response_model=CompareResponse)
def compare_schedules(
    request: CompareRequest, _user: object = Depends(optional_auth)
) -> CompareResponse:
    """Compare two uploaded projects (baseline vs update).

    Args:
        request: Contains baseline_id and update_id.

    Raises:
        HTTPException: If either project is not found.
    """
    store = get_store()

    baseline = store.get(request.baseline_id)
    if baseline is None:
        raise HTTPException(
            status_code=404, detail=f"Baseline project not found: {request.baseline_id}"
        )

    update = store.get(request.update_id)
    if update is None:
        raise HTTPException(
            status_code=404, detail=f"Update project not found: {request.update_id}"
        )

    comparison = ScheduleComparison(baseline, update)
    result = comparison.compare()

    return CompareResponse(
        activities_added=[ActivityChangeSchema(**asdict(c)) for c in result.activities_added],
        activities_deleted=[ActivityChangeSchema(**asdict(c)) for c in result.activities_deleted],
        activity_modifications=[
            ActivityChangeSchema(**asdict(c)) for c in result.activity_modifications
        ],
        duration_changes=[ActivityChangeSchema(**asdict(c)) for c in result.duration_changes],
        relationships_added=[
            RelationshipChangeSchema(**asdict(c)) for c in result.relationships_added
        ],
        relationships_deleted=[
            RelationshipChangeSchema(**asdict(c)) for c in result.relationships_deleted
        ],
        relationships_modified=[
            RelationshipChangeSchema(**asdict(c)) for c in result.relationships_modified
        ],
        significant_float_changes=[
            FloatChangeSchema(**asdict(c)) for c in result.significant_float_changes
        ],
        constraint_changes=[ActivityChangeSchema(**asdict(c)) for c in result.constraint_changes],
        manipulation_flags=[ManipulationFlagSchema(**asdict(c)) for c in result.manipulation_flags],
        code_restructuring=[
            CodeRestructuringSchema(**asdict(c)) for c in result.code_restructuring
        ],
        match_stats=MatchStatsSchema(**asdict(result.match_stats)),
        changed_percentage=result.changed_percentage,
        critical_path_changed=result.critical_path_changed,
        activities_joined_cp=result.activities_joined_cp,
        activities_left_cp=result.activities_left_cp,
        summary=result.summary,
    )


# ------------------------------------------------------------------
# Forensic Analysis (CPA / Window Analysis)
# ------------------------------------------------------------------


def _window_to_schema(wr: Any) -> WindowSchema:
    """Convert a WindowResult dataclass to a WindowSchema.

    Args:
        wr: A ``WindowResult`` instance.

    Returns:
        A ``WindowSchema`` suitable for JSON serialisation.
    """
    return WindowSchema(
        window_number=wr.window.window_number,
        window_id=wr.window.window_id,
        baseline_project_id=wr.window.baseline_project_id,
        update_project_id=wr.window.update_project_id,
        start_date=wr.window.start_date.isoformat() if wr.window.start_date else None,
        end_date=wr.window.end_date.isoformat() if wr.window.end_date else None,
        completion_date_start=(
            wr.completion_date_start.isoformat() if wr.completion_date_start else None
        ),
        completion_date_end=(
            wr.completion_date_end.isoformat() if wr.completion_date_end else None
        ),
        delay_days=wr.delay_days,
        cumulative_delay=wr.cumulative_delay,
        critical_path_start=wr.critical_path_start,
        critical_path_end=wr.critical_path_end,
        cp_activities_joined=wr.cp_activities_joined,
        cp_activities_left=wr.cp_activities_left,
        driving_activity=wr.driving_activity,
        comparison_summary=wr.comparison.summary if wr.comparison else {},
    )


@app.post(
    "/api/v1/forensic/create-timeline",
    response_model=TimelineDetailSchema,
)
def create_timeline(
    request: CreateTimelineRequest, _user: object = Depends(optional_auth)
) -> TimelineDetailSchema:
    """Create a forensic CPA timeline from multiple schedule updates.

    Fetches each referenced project from the store, sorts by data date,
    runs the ``ForensicAnalyzer``, stores the result, and returns the
    full timeline.

    Args:
        request: Contains a list of project_ids (minimum 2).

    Raises:
        HTTPException: If any project is not found or analysis fails.
    """
    store = get_store()
    tl_store = get_timeline_store()

    schedules: list[ParsedSchedule] = []
    for pid in request.project_ids:
        schedule = store.get(pid)
        if schedule is None:
            raise HTTPException(status_code=404, detail=f"Project not found: {pid}")
        schedules.append(schedule)

    try:
        analyzer = ForensicAnalyzer(schedules, list(request.project_ids))
        timeline = analyzer.analyze()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Forensic analysis failed: {exc}")

    tid = tl_store.add(timeline)

    return TimelineDetailSchema(
        timeline_id=tid,
        project_name=timeline.project_name,
        schedule_count=timeline.schedule_count,
        total_delay_days=timeline.total_delay_days,
        contract_completion=(
            timeline.contract_completion.isoformat() if timeline.contract_completion else None
        ),
        current_completion=(
            timeline.current_completion.isoformat() if timeline.current_completion else None
        ),
        windows=[_window_to_schema(w) for w in timeline.windows],
        summary=timeline.summary,
    )


@app.get(
    "/api/v1/forensic/timelines",
    response_model=TimelineListResponse,
)
def list_timelines(_user: object = Depends(optional_auth)) -> TimelineListResponse:
    """List all forensic timelines."""
    tl_store = get_timeline_store()
    items = [TimelineSummarySchema(**t) for t in tl_store.list_all()]
    return TimelineListResponse(timelines=items)


@app.get(
    "/api/v1/forensic/timelines/{timeline_id}",
    response_model=TimelineDetailSchema,
)
def get_timeline(timeline_id: str, _user: object = Depends(optional_auth)) -> TimelineDetailSchema:
    """Get full forensic timeline with all window results.

    Args:
        timeline_id: The stored timeline identifier.

    Raises:
        HTTPException: If the timeline is not found.
    """
    tl_store = get_timeline_store()
    timeline = tl_store.get(timeline_id)
    if timeline is None:
        raise HTTPException(status_code=404, detail="Timeline not found")

    return TimelineDetailSchema(
        timeline_id=timeline.timeline_id,
        project_name=timeline.project_name,
        schedule_count=timeline.schedule_count,
        total_delay_days=timeline.total_delay_days,
        contract_completion=(
            timeline.contract_completion.isoformat() if timeline.contract_completion else None
        ),
        current_completion=(
            timeline.current_completion.isoformat() if timeline.current_completion else None
        ),
        windows=[_window_to_schema(w) for w in timeline.windows],
        summary=timeline.summary,
    )


@app.get(
    "/api/v1/forensic/timelines/{timeline_id}/delay-trend",
    response_model=DelayTrendResponse,
)
def get_delay_trend(timeline_id: str, _user: object = Depends(optional_auth)) -> DelayTrendResponse:
    """Return delay trend data for charting.

    Each point represents one analysis window's data date and the
    forecasted completion date at that point.

    Args:
        timeline_id: The stored timeline identifier.

    Raises:
        HTTPException: If the timeline is not found.
    """
    tl_store = get_timeline_store()
    timeline = tl_store.get(timeline_id)
    if timeline is None:
        raise HTTPException(status_code=404, detail="Timeline not found")

    points: list[DelayTrendPoint] = []
    for wr in timeline.windows:
        points.append(
            DelayTrendPoint(
                window_id=wr.window.window_id,
                window_number=wr.window.window_number,
                data_date=(wr.window.end_date.isoformat() if wr.window.end_date else None),
                completion_date=(
                    wr.completion_date_end.isoformat() if wr.completion_date_end else None
                ),
                delay_days=wr.delay_days,
                cumulative_delay=wr.cumulative_delay,
            )
        )

    return DelayTrendResponse(
        timeline_id=timeline.timeline_id,
        contract_completion=(
            timeline.contract_completion.isoformat() if timeline.contract_completion else None
        ),
        points=points,
    )


# ------------------------------------------------------------------
# TIA (Time Impact Analysis)
# ------------------------------------------------------------------


def _fragment_schema_to_model(schema: DelayFragmentSchema) -> DelayFragment:
    """Convert a DelayFragmentSchema to a DelayFragment dataclass.

    Args:
        schema: The Pydantic schema from the request body.

    Returns:
        A ``DelayFragment`` dataclass instance.
    """
    activities = [
        FragmentActivity(
            fragment_activity_id=a.fragment_activity_id,
            name=a.name,
            duration_hours=a.duration_hours,
            predecessors=a.predecessors,
            successors=a.successors,
        )
        for a in schema.activities
    ]

    return DelayFragment(
        fragment_id=schema.fragment_id,
        name=schema.name,
        description=schema.description,
        responsible_party=ResponsibleParty(schema.responsible_party),
        activities=activities,
    )


def _analysis_to_schema(analysis: Any) -> TIAAnalysisSchema:
    """Convert a TIAAnalysis dataclass to its Pydantic schema.

    Args:
        analysis: A ``TIAAnalysis`` instance.

    Returns:
        A ``TIAAnalysisSchema`` for JSON serialisation.
    """
    fragment_schemas = [
        DelayFragmentSchema(
            fragment_id=f.fragment_id,
            name=f.name,
            description=f.description,
            responsible_party=f.responsible_party.value,
            activities=[
                FragmentActivitySchema(
                    fragment_activity_id=a.fragment_activity_id,
                    name=a.name,
                    duration_hours=a.duration_hours,
                    predecessors=a.predecessors,
                    successors=a.successors,
                )
                for a in f.activities
            ],
        )
        for f in analysis.fragments
    ]

    result_schemas = [
        TIAResultSchema(
            fragment_id=r.fragment.fragment_id,
            fragment_name=r.fragment.name,
            responsible_party=r.fragment.responsible_party.value,
            unimpacted_completion_days=r.unimpacted_completion_days,
            impacted_completion_days=r.impacted_completion_days,
            delay_days=r.delay_days,
            critical_path_affected=r.critical_path_affected,
            float_consumed_hours=r.float_consumed_hours,
            delay_type=r.delay_type.value,
            concurrent_with=r.concurrent_with,
            impacted_driving_path=r.impacted_driving_path,
        )
        for r in analysis.results
    ]

    return TIAAnalysisSchema(
        analysis_id=analysis.analysis_id,
        project_name=analysis.project_name,
        base_project_id=analysis.base_project_id,
        fragments=fragment_schemas,
        results=result_schemas,
        total_owner_delay=analysis.total_owner_delay,
        total_contractor_delay=analysis.total_contractor_delay,
        total_shared_delay=analysis.total_shared_delay,
        net_delay=analysis.net_delay,
        summary=analysis.summary,
    )


@app.post("/api/v1/tia/analyze", response_model=TIAAnalysisSchema)
def tia_analyze(
    request: TIAAnalyzeRequest, _user: object = Depends(optional_auth)
) -> TIAAnalysisSchema:
    """Run Time Impact Analysis on a project with delay fragments.

    Per AACE RP 52R-06, inserts each fragment into the schedule network,
    runs CPM, and measures the impact on the project completion date.

    Args:
        request: Contains project_id and fragment definitions.

    Raises:
        HTTPException: If the project is not found or analysis fails.
    """
    store = get_store()
    tia_store = get_tia_store()

    schedule = store.get(request.project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail=f"Project not found: {request.project_id}")

    # Convert schemas to domain models
    fragments = [_fragment_schema_to_model(f) for f in request.fragments]

    try:
        analyzer = TimeImpactAnalyzer(schedule)
        analysis = analyzer.analyze_all(fragments)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"TIA analysis failed: {exc}")

    tia_store.add(analysis)
    return _analysis_to_schema(analysis)


@app.get("/api/v1/tia/analyses", response_model=TIAListResponse)
def list_tia_analyses(_user: object = Depends(optional_auth)) -> TIAListResponse:
    """List all TIA analyses."""
    tia_store = get_tia_store()
    items = [TIAAnalysisSummarySchema(**a) for a in tia_store.list_all()]
    return TIAListResponse(analyses=items)


@app.get(
    "/api/v1/tia/analyses/{analysis_id}",
    response_model=TIAAnalysisSchema,
)
def get_tia_analysis(analysis_id: str, _user: object = Depends(optional_auth)) -> TIAAnalysisSchema:
    """Get full TIA analysis with all fragment results.

    Args:
        analysis_id: The stored analysis identifier.

    Raises:
        HTTPException: If the analysis is not found.
    """
    tia_store = get_tia_store()
    analysis = tia_store.get(analysis_id)
    if analysis is None:
        raise HTTPException(status_code=404, detail="TIA analysis not found")

    return _analysis_to_schema(analysis)


@app.get(
    "/api/v1/tia/analyses/{analysis_id}/summary",
    response_model=TIASummaryResponse,
)
def get_tia_summary(analysis_id: str, _user: object = Depends(optional_auth)) -> TIASummaryResponse:
    """Get delay-by-responsibility summary for a TIA analysis.

    Args:
        analysis_id: The stored analysis identifier.

    Raises:
        HTTPException: If the analysis is not found.
    """
    tia_store = get_tia_store()
    analysis = tia_store.get(analysis_id)
    if analysis is None:
        raise HTTPException(status_code=404, detail="TIA analysis not found")

    return TIASummaryResponse(
        analysis_id=analysis.analysis_id,
        total_owner_delay=analysis.total_owner_delay,
        total_contractor_delay=analysis.total_contractor_delay,
        total_shared_delay=analysis.total_shared_delay,
        net_delay=analysis.net_delay,
        summary=analysis.summary,
    )


# ------------------------------------------------------------------
# Contract Compliance
# ------------------------------------------------------------------


@app.post("/api/v1/contract/check", response_model=ContractCheckResponse)
def contract_check(
    request: ContractCheckRequest, _user: object = Depends(optional_auth)
) -> ContractCheckResponse:
    """Run contract compliance checks against a TIA analysis.

    Evaluates each fragment against standard construction contract
    provisions per AIA A201, ConsensusDocs 200, and FIDIC conditions.

    Args:
        request: Contains the analysis_id to check.

    Raises:
        HTTPException: If the analysis is not found.
    """
    tia_store = get_tia_store()
    analysis = tia_store.get(request.analysis_id)
    if analysis is None:
        raise HTTPException(
            status_code=404, detail=f"TIA analysis not found: {request.analysis_id}"
        )

    checker = ContractComplianceChecker()
    checks = checker.check_all(analysis.fragments, analysis.results)

    check_schemas = [
        ComplianceCheckSchema(
            fragment_id=c.fragment_id,
            fragment_name=c.fragment_name,
            provision_id=c.provision.provision_id,
            provision_name=c.provision.name,
            provision_category=c.provision.category.value,
            status=c.status.value,
            finding=c.finding,
            recommendation=c.recommendation,
            details=c.details,
        )
        for c in checks
    ]

    warnings = sum(1 for c in checks if c.status.value == "warning")
    failures = sum(1 for c in checks if c.status.value == "fail")

    return ContractCheckResponse(
        analysis_id=request.analysis_id,
        checks=check_schemas,
        total_checks=len(checks),
        warnings=warnings,
        failures=failures,
    )


@app.get(
    "/api/v1/contract/provisions",
    response_model=ContractProvisionsResponse,
)
def list_contract_provisions(_user: object = Depends(optional_auth)) -> ContractProvisionsResponse:
    """List default contract provisions used for compliance checking."""
    checker = ContractComplianceChecker()
    provisions = [
        ContractProvisionSchema(
            provision_id=p.provision_id,
            name=p.name,
            description=p.description,
            category=p.category.value,
            reference=p.reference,
            threshold_days=p.threshold_days,
            details=p.details,
        )
        for p in checker.provisions
    ]
    return ContractProvisionsResponse(provisions=provisions)


# ------------------------------------------------------------------
# EVM (Earned Value Management)
# ------------------------------------------------------------------


def _evm_metrics_to_schema(m: Any) -> EVMMetricsSchema:
    """Convert an EVMMetrics dataclass to its Pydantic schema.

    Args:
        m: An ``EVMMetrics`` instance.

    Returns:
        An ``EVMMetricsSchema`` for JSON serialisation.
    """
    return EVMMetricsSchema(
        scope_name=m.scope_name,
        scope_id=m.scope_id,
        bac=round(m.bac, 2),
        pv=round(m.pv, 2),
        ev=round(m.ev, 2),
        ac=round(m.ac, 2),
        sv=round(m.sv, 2),
        cv=round(m.cv, 2),
        spi=round(m.spi, 3),
        cpi=round(m.cpi, 3),
        eac_cpi=round(m.eac_cpi, 2),
        eac_combined=round(m.eac_combined, 2),
        etc=round(m.etc, 2),
        vac=round(m.vac, 2),
        tcpi=round(m.tcpi, 3),
        percent_complete_ev=round(m.percent_complete_ev, 1),
        percent_spent=round(m.percent_spent, 1),
    )


def _evm_result_to_schema(result: Any, project_id: str = "") -> EVMAnalysisSchema:
    """Convert an EVMAnalysisResult to its Pydantic schema.

    Args:
        result: An ``EVMAnalysisResult`` instance.
        project_id: The project store identifier.

    Returns:
        An ``EVMAnalysisSchema`` for JSON serialisation.
    """
    wbs_schemas = [
        WBSMetricsSchema(
            wbs_id=w.wbs_id,
            wbs_name=w.wbs_name,
            metrics=_evm_metrics_to_schema(w.metrics),
            activity_count=w.activity_count,
        )
        for w in result.wbs_breakdown
    ]

    s_curve_schemas = [
        SCurvePointSchema(
            date=p.date,
            cumulative_pv=p.cumulative_pv,
            cumulative_ev=p.cumulative_ev,
            cumulative_ac=p.cumulative_ac,
        )
        for p in result.s_curve
    ]

    return EVMAnalysisSchema(
        analysis_id=result.analysis_id,
        project_name=result.project_name,
        project_id=project_id,
        data_date=result.data_date,
        metrics=_evm_metrics_to_schema(result.metrics),
        wbs_breakdown=wbs_schemas,
        s_curve=s_curve_schemas,
        schedule_health=HealthClassificationSchema(
            index_name=result.schedule_health.index_name,
            value=result.schedule_health.value,
            status=result.schedule_health.status,
            label=result.schedule_health.label,
        ),
        cost_health=HealthClassificationSchema(
            index_name=result.cost_health.index_name,
            value=result.cost_health.value,
            status=result.cost_health.status,
            label=result.cost_health.label,
        ),
        forecast=result.forecast,
        summary=result.summary,
    )


@app.post("/api/v1/evm/analyze/{project_id}", response_model=EVMAnalysisSchema)
def run_evm_analysis(project_id: str, _user: object = Depends(optional_auth)) -> EVMAnalysisSchema:
    """Run Earned Value Management analysis on a project.

    Computes SPI, CPI, SV, CV, EAC, ETC, VAC, TCPI from resource
    assignment cost data and activity physical percent complete.
    Per ANSI/EIA-748 and AACE RP 10S-90.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found or analysis fails.
    """
    store = get_store()
    evm_store = get_evm_store()

    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail=f"Project not found: {project_id}")

    try:
        analyzer = EVMAnalyzer(schedule)
        result = analyzer.analyze()
        result.project_id = project_id
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"EVM analysis failed: {exc}")

    evm_store.add(result)
    return _evm_result_to_schema(result, project_id)


@app.get("/api/v1/evm/analyses", response_model=EVMListResponse)
def list_evm_analyses(_user: object = Depends(optional_auth)) -> EVMListResponse:
    """List all EVM analyses."""
    evm_store = get_evm_store()
    items = [EVMAnalysisSummarySchema(**a) for a in evm_store.list_all()]
    return EVMListResponse(analyses=items)


@app.get("/api/v1/evm/analyses/{analysis_id}", response_model=EVMAnalysisSchema)
def get_evm_analysis(analysis_id: str, _user: object = Depends(optional_auth)) -> EVMAnalysisSchema:
    """Get full EVM analysis with all metrics.

    Args:
        analysis_id: The stored analysis identifier.

    Raises:
        HTTPException: If the analysis is not found.
    """
    evm_store = get_evm_store()
    result = evm_store.get(analysis_id)
    if result is None:
        raise HTTPException(status_code=404, detail="EVM analysis not found")

    return _evm_result_to_schema(result, result.project_id)


@app.get("/api/v1/evm/analyses/{analysis_id}/s-curve", response_model=SCurveResponse)
def get_evm_s_curve(analysis_id: str, _user: object = Depends(optional_auth)) -> SCurveResponse:
    """Get S-curve data for an EVM analysis.

    Returns time-phased cumulative PV, EV, and AC data points
    suitable for charting.

    Args:
        analysis_id: The stored analysis identifier.

    Raises:
        HTTPException: If the analysis is not found.
    """
    evm_store = get_evm_store()
    result = evm_store.get(analysis_id)
    if result is None:
        raise HTTPException(status_code=404, detail="EVM analysis not found")

    points = [
        SCurvePointSchema(
            date=p.date,
            cumulative_pv=p.cumulative_pv,
            cumulative_ev=p.cumulative_ev,
            cumulative_ac=p.cumulative_ac,
        )
        for p in result.s_curve
    ]

    return SCurveResponse(analysis_id=analysis_id, points=points)


@app.get(
    "/api/v1/evm/analyses/{analysis_id}/wbs-drill",
    response_model=WBSDrillResponse,
)
def get_evm_wbs_drill(analysis_id: str, _user: object = Depends(optional_auth)) -> WBSDrillResponse:
    """Get WBS-level EVM breakdown for an analysis.

    Args:
        analysis_id: The stored analysis identifier.

    Raises:
        HTTPException: If the analysis is not found.
    """
    evm_store = get_evm_store()
    result = evm_store.get(analysis_id)
    if result is None:
        raise HTTPException(status_code=404, detail="EVM analysis not found")

    wbs_schemas = [
        WBSMetricsSchema(
            wbs_id=w.wbs_id,
            wbs_name=w.wbs_name,
            metrics=_evm_metrics_to_schema(w.metrics),
            activity_count=w.activity_count,
        )
        for w in result.wbs_breakdown
    ]

    return WBSDrillResponse(analysis_id=analysis_id, wbs_breakdown=wbs_schemas)


@app.get(
    "/api/v1/evm/analyses/{analysis_id}/forecast",
    response_model=ForecastResponse,
)
def get_evm_forecast(analysis_id: str, _user: object = Depends(optional_auth)) -> ForecastResponse:
    """Get EAC scenario forecasts for an analysis.

    Returns multiple Estimate at Completion scenarios:
    - EAC (CPI): BAC / CPI
    - EAC (Combined): AC + (BAC - EV) / (CPI * SPI)
    - EAC (New ETC): AC + (BAC - EV)
    - ETC, VAC, TCPI

    Args:
        analysis_id: The stored analysis identifier.

    Raises:
        HTTPException: If the analysis is not found.
    """
    evm_store = get_evm_store()
    result = evm_store.get(analysis_id)
    if result is None:
        raise HTTPException(status_code=404, detail="EVM analysis not found")

    return ForecastResponse(
        analysis_id=analysis_id,
        **result.forecast,
    )


# ------------------------------------------------------------------
# Risk (Monte Carlo QSRA)
# ------------------------------------------------------------------


def _simulation_to_schema(result: Any) -> SimulationResultSchema:
    """Convert a SimulationResult dataclass to its Pydantic schema.

    Args:
        result: A ``SimulationResult`` instance.

    Returns:
        A ``SimulationResultSchema`` for JSON serialisation.
    """
    return SimulationResultSchema(
        simulation_id=result.simulation_id,
        project_name=result.project_name,
        project_id=result.project_id,
        iterations=result.iterations,
        deterministic_days=result.deterministic_days,
        mean_days=result.mean_days,
        std_days=result.std_days,
        p_values=[
            PValueSchema(
                percentile=pv.percentile,
                duration_days=pv.duration_days,
                delta_days=pv.delta_days,
            )
            for pv in result.p_values
        ],
        histogram=[
            HistogramBinSchema(
                bin_start=b.bin_start,
                bin_end=b.bin_end,
                count=b.count,
                frequency=b.frequency,
            )
            for b in result.histogram
        ],
        criticality=[
            CriticalityEntrySchema(
                activity_id=c.activity_id,
                activity_name=c.activity_name,
                criticality_pct=c.criticality_pct,
            )
            for c in result.criticality
        ],
        sensitivity=[
            SensitivityEntrySchema(
                activity_id=s.activity_id,
                activity_name=s.activity_name,
                correlation=s.correlation,
            )
            for s in result.sensitivity
        ],
        s_curve=[
            RiskSCurvePointSchema(
                duration_days=p.duration_days,
                cumulative_probability=p.cumulative_probability,
            )
            for p in result.s_curve
        ],
    )


@app.post(
    "/api/v1/risk/simulate/{project_id}",
    response_model=SimulationResultSchema,
)
def run_risk_simulation(
    project_id: str,
    request: RunSimulationRequest,
    _user: object = Depends(optional_auth),
) -> SimulationResultSchema:
    """Run Monte Carlo schedule risk simulation (QSRA) on a project.

    Performs *N* iterations sampling activity durations from probability
    distributions, running CPM each time, and computing completion date
    probability distributions, criticality indices, and sensitivity.
    Per AACE RP 57R-09.

    Args:
        project_id: The stored project identifier.
        request: Simulation configuration, duration risks, and risk events.

    Raises:
        HTTPException: If the project is not found or simulation fails.
    """
    store = get_store()
    risk_store = get_risk_store()

    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail=f"Project not found: {project_id}")

    # Convert request schemas to domain models
    config = None
    if request.config:
        config = SimulationConfig(
            iterations=request.config.iterations,
            default_distribution=DistributionType(request.config.default_distribution),
            default_uncertainty=request.config.default_uncertainty,
            seed=request.config.seed,
            confidence_levels=request.config.confidence_levels,
        )

    duration_risks = (
        [
            DurationRisk(
                activity_id=r.activity_id,
                distribution=DistributionType(r.distribution),
                min_duration=r.min_duration,
                most_likely=r.most_likely,
                max_duration=r.max_duration,
            )
            for r in request.duration_risks
        ]
        if request.duration_risks
        else None
    )

    risk_events = (
        [
            RiskEvent(
                risk_id=e.risk_id,
                name=e.name,
                probability=e.probability,
                impact_hours=e.impact_hours,
                affected_activities=e.affected_activities,
            )
            for e in request.risk_events
        ]
        if request.risk_events
        else None
    )

    try:
        simulator = MonteCarloSimulator(schedule, config)
        result = simulator.simulate(duration_risks, risk_events)
        result.project_id = project_id
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Risk simulation failed: {exc}")

    risk_store.add(result)
    return _simulation_to_schema(result)


@app.get("/api/v1/risk/simulations", response_model=SimulationListResponse)
def list_risk_simulations(_user: object = Depends(optional_auth)) -> SimulationListResponse:
    """List all risk simulations."""
    risk_store = get_risk_store()
    items = [SimulationSummarySchema(**s) for s in risk_store.list_all()]
    return SimulationListResponse(simulations=items)


@app.get(
    "/api/v1/risk/simulations/{simulation_id}",
    response_model=SimulationResultSchema,
)
def get_risk_simulation(
    simulation_id: str, _user: object = Depends(optional_auth)
) -> SimulationResultSchema:
    """Get full risk simulation result with all analysis data.

    Args:
        simulation_id: The stored simulation identifier.

    Raises:
        HTTPException: If the simulation is not found.
    """
    risk_store = get_risk_store()
    result = risk_store.get(simulation_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Risk simulation not found")

    return _simulation_to_schema(result)


@app.get(
    "/api/v1/risk/simulations/{simulation_id}/histogram",
    response_model=HistogramResponse,
)
def get_risk_histogram(
    simulation_id: str, _user: object = Depends(optional_auth)
) -> HistogramResponse:
    """Get histogram data for a risk simulation.

    Returns bin counts and P-value overlay lines suitable for
    charting a probability distribution of project completion.

    Args:
        simulation_id: The stored simulation identifier.

    Raises:
        HTTPException: If the simulation is not found.
    """
    risk_store = get_risk_store()
    result = risk_store.get(simulation_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Risk simulation not found")

    return HistogramResponse(
        simulation_id=simulation_id,
        deterministic_days=result.deterministic_days,
        bins=[
            HistogramBinSchema(
                bin_start=b.bin_start,
                bin_end=b.bin_end,
                count=b.count,
                frequency=b.frequency,
            )
            for b in result.histogram
        ],
        p_values=[
            PValueSchema(
                percentile=pv.percentile,
                duration_days=pv.duration_days,
                delta_days=pv.delta_days,
            )
            for pv in result.p_values
        ],
    )


@app.get(
    "/api/v1/risk/simulations/{simulation_id}/tornado",
    response_model=TornadoResponse,
)
def get_risk_tornado(simulation_id: str, _user: object = Depends(optional_auth)) -> TornadoResponse:
    """Get sensitivity / tornado data for a risk simulation.

    Returns Spearman rank correlations between each activity's
    sampled duration and the project completion date, sorted by
    absolute correlation (top 15).

    Args:
        simulation_id: The stored simulation identifier.

    Raises:
        HTTPException: If the simulation is not found.
    """
    risk_store = get_risk_store()
    result = risk_store.get(simulation_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Risk simulation not found")

    # Return top 15 by absolute correlation
    top_entries = result.sensitivity[:15]

    return TornadoResponse(
        simulation_id=simulation_id,
        entries=[
            SensitivityEntrySchema(
                activity_id=s.activity_id,
                activity_name=s.activity_name,
                correlation=s.correlation,
            )
            for s in top_entries
        ],
    )


@app.get(
    "/api/v1/risk/simulations/{simulation_id}/criticality",
    response_model=CriticalityResponse,
)
def get_risk_criticality(
    simulation_id: str, _user: object = Depends(optional_auth)
) -> CriticalityResponse:
    """Get criticality index data for a risk simulation.

    Returns the percentage of iterations in which each activity
    appeared on the critical path.

    Args:
        simulation_id: The stored simulation identifier.

    Raises:
        HTTPException: If the simulation is not found.
    """
    risk_store = get_risk_store()
    result = risk_store.get(simulation_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Risk simulation not found")

    return CriticalityResponse(
        simulation_id=simulation_id,
        entries=[
            CriticalityEntrySchema(
                activity_id=c.activity_id,
                activity_name=c.activity_name,
                criticality_pct=c.criticality_pct,
            )
            for c in result.criticality
        ],
    )


@app.get(
    "/api/v1/risk/simulations/{simulation_id}/s-curve",
    response_model=RiskSCurveResponse,
)
def get_risk_s_curve(
    simulation_id: str, _user: object = Depends(optional_auth)
) -> RiskSCurveResponse:
    """Get cumulative probability S-curve data for a risk simulation.

    Returns sorted completion durations with their cumulative
    probability, suitable for charting a sigmoid/S-curve.

    Args:
        simulation_id: The stored simulation identifier.

    Raises:
        HTTPException: If the simulation is not found.
    """
    risk_store = get_risk_store()
    result = risk_store.get(simulation_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Risk simulation not found")

    return RiskSCurveResponse(
        simulation_id=simulation_id,
        deterministic_days=result.deterministic_days,
        points=[
            RiskSCurvePointSchema(
                duration_days=p.duration_days,
                cumulative_probability=p.cumulative_probability,
            )
            for p in result.s_curve
        ],
        p_values=[
            PValueSchema(
                percentile=pv.percentile,
                duration_days=pv.duration_days,
                delta_days=pv.delta_days,
            )
            for pv in result.p_values
        ],
    )


# ══════════════════════════════════════════════════════════
# Intelligence v0.8 — Health Score, Float Trends, Alerts, Dashboard
# ══════════════════════════════════════════════════════════


@app.get(
    "/api/v1/projects/{project_id}/health",
    response_model=ScheduleHealthResponse,
)
def get_project_health(
    project_id: str,
    baseline_id: str | None = None,
    _user: object = Depends(optional_auth),
) -> ScheduleHealthResponse:
    """Get the composite schedule health score for a project.

    Computes a 0-100 score combining DCMA structural quality, float
    health, logic integrity, and trend direction.  If ``baseline_id``
    is provided, the trend component uses the baseline for comparison.

    Args:
        project_id: The stored project identifier.
        baseline_id: Optional baseline project for trend analysis.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    baseline = None
    if baseline_id:
        baseline = store.get(baseline_id)
        if baseline is None:
            raise HTTPException(status_code=404, detail="Baseline project not found")

    calc = HealthScoreCalculator(schedule, baseline=baseline)
    score = calc.calculate()

    return ScheduleHealthResponse(
        overall=score.overall,
        dcma_component=score.dcma_component,
        float_component=score.float_component,
        logic_component=score.logic_component,
        trend_component=score.trend_component,
        dcma_raw=score.dcma_raw,
        float_raw=score.float_raw,
        logic_raw=score.logic_raw,
        trend_raw=score.trend_raw,
        rating=score.rating,
        trend_arrow=score.trend_arrow,
        details=score.details,
    )


@app.get(
    "/api/v1/projects/{project_id}/float-trends",
    response_model=FloatTrendResponse,
)
def get_float_trends(
    project_id: str,
    baseline_id: str | None = None,
    _user: object = Depends(optional_auth),
) -> FloatTrendResponse:
    """Get float trend data between a baseline and update schedule.

    Computes Float Erosion Index, Near-Critical Drift, CP Stability,
    and per-activity float deltas.

    Args:
        project_id: The update project identifier.
        baseline_id: The baseline project identifier.

    Raises:
        HTTPException: If projects are not found.
    """
    store = get_store()
    update = store.get(project_id)
    if update is None:
        raise HTTPException(status_code=404, detail="Project not found")

    if not baseline_id:
        raise HTTPException(
            status_code=400,
            detail="baseline_id query parameter is required for float trend analysis",
        )

    baseline = store.get(baseline_id)
    if baseline is None:
        raise HTTPException(status_code=404, detail="Baseline project not found")

    analyzer = FloatTrendAnalyzer(baseline, update)
    result = analyzer.analyze()

    return FloatTrendResponse(
        fei=result.fei,
        near_critical_drift=result.near_critical_drift,
        cp_stability=result.cp_stability,
        activity_trends=[
            ActivityFloatTrendSchema(
                task_code=t.task_code,
                task_name=t.task_name,
                wbs_id=t.wbs_id,
                old_float_days=t.old_float_days,
                new_float_days=t.new_float_days,
                delta_days=t.delta_days,
                direction=t.direction,
                is_critical_baseline=t.is_critical_baseline,
                is_critical_update=t.is_critical_update,
                progress_pct=t.progress_pct,
            )
            for t in result.activity_trends
        ],
        wbs_velocity=result.wbs_velocity,
        thresholds=result.thresholds,
        days_between_updates=result.days_between_updates,
        total_matched=result.total_matched,
        summary=result.summary,
    )


@app.get("/api/v1/projects/{project_id}/root-cause")
def get_root_cause(
    project_id: str,
    activity_id: str | None = None,
    _user: object = Depends(optional_auth),
) -> dict:
    """Trace backwards through the dependency network to find the root cause.

    Starting from a target activity (or the project completion driver if
    not specified), walks backwards through driving predecessors to
    identify the originating delay event.

    Args:
        project_id: The stored project identifier.
        activity_id: Optional target activity ID. If omitted, uses the
            activity with the latest early finish.

    Returns:
        RootCauseResult as dict with the driving chain.

    Raises:
        HTTPException: If the project is not found.

    References:
        AACE RP 49R-06 — Identifying Critical Activities.
        AACE RP 29R-03 — Forensic Schedule Analysis.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    from src.analytics.root_cause import analyze_root_cause

    result = analyze_root_cause(schedule, target_task_id=activity_id)
    return asdict(result)


@app.post("/api/v1/projects/{project_id}/ask")
async def ask_schedule(
    project_id: str,
    body: dict,
    _user: object = Depends(optional_auth),
) -> dict:
    """Ask a natural language question about a schedule.

    Uses Claude API to interpret the question and generate an answer
    grounded in the schedule's actual data. Does NOT send raw schedule
    data to the API — only a compact statistical summary.

    Args:
        project_id: The stored project identifier.
        body: JSON with ``question`` (required) and optional ``api_key``.

    Returns:
        Dict with ``answer``, ``question``, ``tokens_used``, ``model``.

    Raises:
        HTTPException: If project not found or API key missing.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    question = body.get("question", "").strip()
    if not question:
        raise HTTPException(status_code=400, detail="question is required")

    api_key = body.get("api_key") or os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(
            status_code=400,
            detail="Anthropic API key required. Pass api_key in body or set ANTHROPIC_API_KEY env var.",
        )

    from src.analytics.nlp_query import query_schedule

    try:
        result = await query_schedule(schedule, question, api_key=api_key)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

    return {
        "question": result.question,
        "answer": result.answer,
        "model": result.model,
        "tokens_used": result.tokens_used,
    }


@app.get("/api/v1/projects/{project_id}/anomalies")
def get_anomalies(
    project_id: str,
    _user: object = Depends(optional_auth),
) -> dict:
    """Detect statistical anomalies in schedule data.

    Uses IQR and z-score methods to flag activities with unusual
    duration, float, progress, or relationship patterns.

    Args:
        project_id: The stored project identifier.

    Returns:
        AnomalyDetectionResult with anomalies sorted by severity.

    References:
        Tukey (1977) — Exploratory Data Analysis (IQR method).
        DCMA 14-Point Assessment — duration and float thresholds.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    from src.analytics.anomaly_detection import detect_anomalies

    result = detect_anomalies(schedule)
    return asdict(result)


@app.get("/api/v1/projects/{project_id}/float-entropy")
def get_float_entropy(
    project_id: str,
    _user: object = Depends(optional_auth),
) -> dict:
    """Compute Shannon entropy of float distribution.

    Measures how uniformly total float is spread across predefined
    buckets.  Low entropy indicates float concentrated in few categories
    (potentially unreliable); high entropy indicates even distribution.

    No baseline required — operates on a single schedule snapshot.

    Args:
        project_id: The stored project identifier.

    Returns:
        FloatEntropyResult as dict with entropy, normalised_entropy,
        distribution, and interpretation.

    Raises:
        HTTPException: If the project is not found.

    References:
        Shannon (1948) — A Mathematical Theory of Communication.
        AACE RP 49R-06 — Identifying Critical Activities.
    """
    store = get_store()
    schedule = store.get(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    result = compute_float_entropy(schedule)
    return asdict(result)


@app.get("/api/v1/projects/{project_id}/constraint-accumulation")
def get_constraint_accumulation(
    project_id: str,
    baseline_id: str | None = None,
    _user: object = Depends(optional_auth),
) -> dict:
    """Compute constraint accumulation rate between two schedule versions.

    Measures the rate at which hard constraints are being added.
    Excessive constraint growth is a schedule manipulation indicator
    per DCMA check #10 and AACE RP 29R-03.

    Args:
        project_id: The update project identifier.
        baseline_id: The baseline project identifier (required).

    Returns:
        ConstraintAccumulationResult as dict.

    Raises:
        HTTPException: If projects are not found or baseline_id missing.

    References:
        DCMA 14-Point Assessment — Check #10 (Hard Constraints).
        AACE RP 29R-03 — Schedule manipulation detection.
    """
    store = get_store()
    update = store.get(project_id)
    if update is None:
        raise HTTPException(status_code=404, detail="Project not found")

    if not baseline_id:
        raise HTTPException(
            status_code=400,
            detail="baseline_id query parameter is required for constraint accumulation",
        )

    baseline = store.get(baseline_id)
    if baseline is None:
        raise HTTPException(status_code=404, detail="Baseline project not found")

    result = compute_constraint_accumulation(baseline, update)
    return asdict(result)


@app.get(
    "/api/v1/projects/{project_id}/alerts",
    response_model=AlertsResponse,
)
def get_project_alerts(
    project_id: str,
    baseline_id: str | None = None,
    _user: object = Depends(optional_auth),
) -> AlertsResponse:
    """Get early warning alerts for a project.

    Runs the 12-rule early warning engine comparing baseline and update
    schedules.  Produces prioritized alerts ranked by severity, confidence,
    and projected impact.

    Args:
        project_id: The update project identifier.
        baseline_id: The baseline project identifier.

    Raises:
        HTTPException: If projects are not found.
    """
    store = get_store()
    update = store.get(project_id)
    if update is None:
        raise HTTPException(status_code=404, detail="Project not found")

    if not baseline_id:
        raise HTTPException(
            status_code=400,
            detail="baseline_id query parameter is required for early warning analysis",
        )

    baseline = store.get(baseline_id)
    if baseline is None:
        raise HTTPException(status_code=404, detail="Baseline project not found")

    engine = EarlyWarningEngine(baseline, update)
    result = engine.analyze()

    return AlertsResponse(
        alerts=[
            AlertSchema(
                rule_id=a.rule_id,
                severity=a.severity,
                title=a.title,
                description=a.description,
                affected_activities=a.affected_activities,
                projected_impact_days=a.projected_impact_days,
                confidence=a.confidence,
                alert_score=a.alert_score,
            )
            for a in result.alerts
        ],
        total_alerts=result.total_alerts,
        critical_count=result.critical_count,
        warning_count=result.warning_count,
        info_count=result.info_count,
        aggregate_score=result.aggregate_score,
        summary=result.summary,
    )


@app.get(
    "/api/v1/dashboard",
    response_model=DashboardKPIs,
)
def get_dashboard(_user: object = Depends(optional_auth)) -> DashboardKPIs:
    """Get portfolio-level dashboard KPIs.

    Returns total projects, average health score, active alerts count,
    and identifies the most critical project.  Computes health scores
    on-the-fly for all loaded projects.
    """
    store = get_store()
    project_ids = store.list_ids()

    if not project_ids:
        return DashboardKPIs()

    health_scores: dict[str, float] = {}

    for pid in project_ids:
        schedule = store.get(pid)
        if schedule is None:
            continue

        try:
            calc = HealthScoreCalculator(schedule)
            score = calc.calculate()
            health_scores[pid] = score.overall
        except Exception:
            health_scores[pid] = 50.0

    most_critical_id = min(health_scores, key=health_scores.get) if health_scores else None
    most_critical_score = health_scores.get(most_critical_id, None) if most_critical_id else None

    avg_health = sum(health_scores.values()) / len(health_scores) if health_scores else 0.0

    return DashboardKPIs(
        total_projects=len(project_ids),
        active_alerts=0,
        avg_health_score=round(avg_health, 1),
        projects_trending_up=0,
        projects_trending_down=0,
        most_critical_project=most_critical_id,
        most_critical_score=most_critical_score,
    )


# ══════════════════════════════════════════════════════════
# PDF Report Generation
# ══════════════════════════════════════════════════════════

_VALID_REPORT_TYPES = {"health", "comparison", "forensic", "tia", "risk", "monthly_review"}


@app.post(
    "/api/v1/reports/generate",
    response_model=GenerateReportResponse,
)
def generate_report(
    request: GenerateReportRequest,
    _user: object = Depends(optional_auth),
) -> GenerateReportResponse:
    """Generate a PDF report. Returns report ID for download.

    Supports 6 report types:
    - health: Schedule Health Report (DCMA + health score + alerts)
    - comparison: Comparison Report (requires baseline_id)
    - forensic: Forensic Report (requires baseline_id for timeline)
    - tia: TIA Report (requires baseline_id)
    - risk: Risk Report (requires baseline_id)
    - monthly_review: Monthly Review Report (health + comparison + alerts)

    Args:
        request: Report generation parameters.

    Raises:
        HTTPException: If the project is not found or report type is invalid.
    """
    if request.report_type not in _VALID_REPORT_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid report_type '{request.report_type}'. "
            f"Valid types: {', '.join(sorted(_VALID_REPORT_TYPES))}",
        )

    store = get_store()
    schedule = store.get(request.project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    generator = ReportGenerator()
    report_type = request.report_type

    try:
        if report_type == "health":
            pdf_bytes = _generate_health_report(generator, schedule, request)
        elif report_type == "comparison":
            pdf_bytes = _generate_comparison_report(generator, schedule, request, store)
        elif report_type == "forensic":
            pdf_bytes = _generate_forensic_report(generator, schedule, request, store)
        elif report_type == "tia":
            pdf_bytes = _generate_tia_report(generator, schedule, request, store)
        elif report_type == "risk":
            pdf_bytes = _generate_risk_report(generator, schedule, request, store)
        elif report_type == "monthly_review":
            pdf_bytes = _generate_monthly_review_report(generator, schedule, request, store)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported report type: {report_type}")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Report generation failed: {exc}")

    from datetime import datetime as _dt

    generated_at = _dt.now().isoformat()

    report_store = get_report_store()
    report_id = report_store.add(
        pdf_bytes,
        {
            "report_type": report_type,
            "project_id": request.project_id,
            "generated_at": generated_at,
        },
    )

    return GenerateReportResponse(
        report_id=report_id,
        report_type=report_type,
        project_id=request.project_id,
        generated_at=generated_at,
    )


def _generate_health_report(
    generator: ReportGenerator,
    schedule: ParsedSchedule,
    request: GenerateReportRequest,
) -> bytes:
    """Generate a health report PDF."""
    from src.analytics.dcma14 import DCMA14Analyzer
    from src.analytics.health_score import HealthScoreCalculator

    dcma = DCMA14Analyzer(schedule)
    dcma_result = dcma.analyze()

    baseline = None
    if request.baseline_id:
        baseline = get_store().get(request.baseline_id)

    calc = HealthScoreCalculator(schedule, baseline=baseline)
    health = calc.calculate()

    alerts_result = None
    if baseline:
        try:
            engine = EarlyWarningEngine(baseline, schedule)
            alerts_result = engine.analyze()
        except Exception:
            pass

    return generator.generate_health_report(schedule, dcma_result, health, alerts_result)


def _generate_comparison_report(
    generator: ReportGenerator,
    schedule: ParsedSchedule,
    request: GenerateReportRequest,
    store: ProjectStore,
) -> bytes:
    """Generate a comparison report PDF."""
    if not request.baseline_id:
        raise HTTPException(
            status_code=400,
            detail="baseline_id is required for comparison reports",
        )

    baseline = store.get(request.baseline_id)
    if baseline is None:
        raise HTTPException(status_code=404, detail="Baseline project not found")

    comparison = ScheduleComparison(baseline, schedule)
    result = comparison.compare()

    return generator.generate_comparison_report(baseline, schedule, result)


def _generate_forensic_report(
    generator: ReportGenerator,
    schedule: ParsedSchedule,
    request: GenerateReportRequest,
    store: ProjectStore,
) -> bytes:
    """Generate a forensic report PDF."""
    if not request.baseline_id:
        raise HTTPException(
            status_code=400,
            detail="baseline_id is required for forensic reports",
        )

    baseline = store.get(request.baseline_id)
    if baseline is None:
        raise HTTPException(status_code=404, detail="Baseline project not found")

    analyzer = ForensicAnalyzer(
        schedules=[baseline, schedule],
        project_ids=[request.baseline_id, request.project_id],
    )
    timeline = analyzer.analyze()

    return generator.generate_forensic_report(timeline)


def _generate_tia_report(
    generator: ReportGenerator,
    schedule: ParsedSchedule,
    request: GenerateReportRequest,
    store: ProjectStore,
) -> bytes:
    """Generate a TIA report PDF."""
    # Look for existing TIA analysis, or create a minimal one
    tia_store = get_tia_store()
    analyses = tia_store.list_all()
    if analyses:
        # Use the most recent TIA analysis
        latest = analyses[-1]
        analysis = tia_store.get(latest["analysis_id"])
        if analysis:
            return generator.generate_tia_report(analysis)

    raise HTTPException(
        status_code=400,
        detail="No TIA analysis found. Please run a TIA analysis first.",
    )


def _generate_risk_report(
    generator: ReportGenerator,
    schedule: ParsedSchedule,
    request: GenerateReportRequest,
    store: ProjectStore,
) -> bytes:
    """Generate a risk report PDF."""
    risk_store = get_risk_store()
    simulations = risk_store.list_all()
    if simulations:
        latest = simulations[-1]
        result = risk_store.get(latest["simulation_id"])
        if result:
            return generator.generate_risk_report(result)

    raise HTTPException(
        status_code=400,
        detail="No risk simulation found. Please run a Monte Carlo simulation first.",
    )


def _generate_monthly_review_report(
    generator: ReportGenerator,
    schedule: ParsedSchedule,
    request: GenerateReportRequest,
    store: ProjectStore,
) -> bytes:
    """Generate a monthly review report PDF.

    Combines health score, DCMA, comparison delta, and early warning alerts
    into a single standardised monthly update report per PMI PMBOK 7 S4.6
    and CMAA (2019) S7.
    """
    from src.analytics.dcma14 import DCMA14Analyzer
    from src.analytics.health_score import HealthScoreCalculator

    dcma = DCMA14Analyzer(schedule)
    dcma_result = dcma.analyze()

    baseline = None
    if request.baseline_id:
        baseline = store.get(request.baseline_id)

    calc = HealthScoreCalculator(schedule, baseline=baseline)
    health = calc.calculate()

    # Comparison with baseline (if provided)
    comparison_result = None
    if baseline:
        comparison = ScheduleComparison(baseline, schedule)
        comparison_result = comparison.compare()

    # Early warning alerts (if baseline provided)
    alerts_result = None
    if baseline:
        try:
            engine = EarlyWarningEngine(baseline, schedule)
            alerts_result = engine.analyze()
        except Exception:
            pass

    return generator.generate_monthly_review_report(
        schedule,
        dcma_result,
        health,
        comparison_result,
        alerts_result,
        baseline,
    )


@app.get("/api/v1/reports/{report_id}/download")
def download_report(
    report_id: str,
    _user: object = Depends(optional_auth),
) -> StreamingResponse:
    """Download a generated PDF report.

    Args:
        report_id: The report identifier from generate_report.

    Raises:
        HTTPException: If the report is not found.
    """
    report_store = get_report_store()
    report = report_store.get(report_id)
    if report is None:
        raise HTTPException(status_code=404, detail="Report not found")

    pdf_bytes = report["bytes"]
    report_type = report.get("report_type", "report")

    # Determine content type: PDF if starts with %PDF, otherwise HTML
    content_type = "application/pdf"
    extension = "pdf"
    if not pdf_bytes[:5].startswith(b"%PDF"):
        content_type = "text/html"
        extension = "html"

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type=content_type,
        headers={
            "Content-Disposition": f'attachment; filename="meridianiq-{report_type}-{report_id}.{extension}"',
        },
    )


# ══════════════════════════════════════════════════════════
# P3 — Report Availability
# ══════════════════════════════════════════════════════════


@app.get("/api/v1/projects/{project_id}/available-reports")
def get_available_reports(
    project_id: str,
    _user: object = Depends(optional_auth),
) -> dict:
    """Check which report types have data available for a project.

    Returns a list of report descriptors with ``ready`` boolean and a
    human-readable ``reason`` when the report is not yet available.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    reports = []

    # --- health: always computable from the schedule itself ---
    reports.append(
        {
            "type": "health",
            "name": "Schedule Health Report",
            "ready": True,
            "reason": "",
        }
    )

    # --- dcma: always computable from the schedule itself ---
    reports.append(
        {
            "type": "dcma",
            "name": "DCMA 14-Point Report",
            "ready": True,
            "reason": "",
        }
    )

    # --- comparison: needs a second upload in the same program ---
    comparison_ready = False
    comparison_reason = "Requires at least two schedule revisions in the same program"
    if hasattr(store, "_upload_program"):
        # InMemoryStore: look for a sibling upload
        program_id = store._upload_program.get(project_id)
        if program_id:
            siblings = [
                pid
                for pid, prog in store._upload_program.items()
                if prog == program_id and pid != project_id
            ]
            if siblings:
                comparison_ready = True
                comparison_reason = ""
    elif hasattr(store, "get_program_revisions"):
        # SupabaseStore: check via program revisions
        programs = store.get_programs(user_id=user_id)
        for prog in programs:
            revisions = store.get_program_revisions(prog["id"], user_id=user_id)
            ids = [r["id"] for r in revisions]
            if project_id in ids and len(ids) >= 2:
                comparison_ready = True
                comparison_reason = ""
                break

    reports.append(
        {
            "type": "comparison",
            "name": "Schedule Comparison Report",
            "ready": comparison_ready,
            "reason": comparison_reason,
        }
    )

    # --- evm: check if an EVM analysis exists for this project ---
    evm_ready = False
    evm_reason = "Run EVM analysis first"
    evm_store = get_evm_store()
    for entry in evm_store.list_all():
        if entry.get("project_id") == project_id:
            evm_ready = True
            evm_reason = ""
            break

    reports.append(
        {
            "type": "evm",
            "name": "Earned Value Report",
            "ready": evm_ready,
            "reason": evm_reason,
        }
    )

    # --- risk: check if a risk simulation exists for this project ---
    risk_ready = False
    risk_reason = "Run Monte Carlo simulation first"
    risk_store = get_risk_store()
    for entry in risk_store.list_all():
        if entry.get("project_id") == project_id:
            risk_ready = True
            risk_reason = ""
            break

    reports.append(
        {
            "type": "risk",
            "name": "Risk / QSRA Report",
            "ready": risk_ready,
            "reason": risk_reason,
        }
    )

    # --- monthly_review: always computable (enhanced with comparison if available) ---
    reports.append(
        {
            "type": "monthly_review",
            "name": "Monthly Review Report",
            "ready": True,
            "reason": "",
        }
    )

    return {"project_id": project_id, "reports": reports}


# ══════════════════════════════════════════════════════════
# P4 — Activity Search (for TIA autocomplete)
# ══════════════════════════════════════════════════════════


@app.get("/api/v1/projects/{project_id}/activities")
def search_activities(
    project_id: str,
    q: str = "",
    limit: int = 20,
    _user: object = Depends(optional_auth),
) -> dict:
    """Search activities by ID or name.  Used by the TIA activity picker.

    Args:
        project_id: The stored project identifier.
        q: Optional search query matched against task_code and task_name.
        limit: Maximum number of results to return (default 20).

    Returns:
        Dict with ``activities`` list and ``total`` count of all activities.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        # Fallback: try get_parsed_schedule (SupabaseStore path)
        if hasattr(store, "get_parsed_schedule"):
            schedule = store.get_parsed_schedule(project_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    total = len(schedule.activities)
    activities: list[dict] = []
    q_lower = q.strip().lower()

    for task in schedule.activities:
        entry = {
            "task_code": task.task_code,
            "task_name": task.task_name,
            "task_type": task.task_type,
            "wbs_id": task.wbs_id,
            "status_code": task.status_code,
        }
        if q_lower:
            searchable = f"{task.task_code} {task.task_name}".lower()
            if q_lower not in searchable:
                continue
        activities.append(entry)
        if len(activities) >= limit:
            break

    return {"activities": activities, "total": total}


# ══════════════════════════════════════════════════════════
# Excel Export
# ══════════════════════════════════════════════════════════


@app.get("/api/v1/projects/{project_id}/export/excel")
def export_excel(
    project_id: str,
    _user: object = Depends(optional_auth),
):
    """Export project schedule data as an Excel workbook.

    Includes sheets for: Activities, Relationships, WBS, and Summary.
    PM and Owner personas use this for custom reporting in Excel.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found or openpyxl is unavailable.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(
            status_code=501, detail="Excel export requires openpyxl (pip install openpyxl)"
        )

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["MeridianIQ Schedule Export"])
    ws_summary.append(
        ["Project", schedule.projects[0].proj_short_name if schedule.projects else ""]
    )
    ws_summary.append(["Activities", len(schedule.activities)])
    ws_summary.append(["Relationships", len(schedule.relationships)])
    ws_summary.append(["WBS Elements", len(schedule.wbs_nodes)])
    ws_summary.append(["Calendars", len(schedule.calendars)])
    ws_summary.append(["Parser Version", schedule.parser_version])

    # Activities sheet
    ws_act = wb.create_sheet("Activities")
    ws_act.append(
        [
            "Task ID",
            "Task Code",
            "Task Name",
            "Type",
            "Status",
            "Total Float (hrs)",
            "Remaining Duration (hrs)",
            "Original Duration (hrs)",
            "Actual Start",
            "Actual Finish",
            "Early Start",
            "Early Finish",
        ]
    )
    for t in schedule.activities:
        ws_act.append(
            [
                t.task_id,
                t.task_code,
                t.task_name,
                t.task_type,
                t.status_code,
                t.total_float_hr_cnt,
                t.remain_drtn_hr_cnt,
                t.target_drtn_hr_cnt,
                str(t.act_start_date) if t.act_start_date else "",
                str(t.act_end_date) if t.act_end_date else "",
                str(t.early_start_date) if t.early_start_date else "",
                str(t.early_end_date) if t.early_end_date else "",
            ]
        )

    # Relationships sheet
    ws_rel = wb.create_sheet("Relationships")
    ws_rel.append(["Task ID", "Predecessor ID", "Type", "Lag (hrs)"])
    for r in schedule.relationships:
        ws_rel.append([r.task_id, r.pred_task_id, r.pred_type, r.lag_hr_cnt])

    # WBS sheet
    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.append(["WBS ID", "Parent WBS ID", "Short Name", "Name"])
    for w in schedule.wbs_nodes:
        ws_wbs.append([w.wbs_id, w.parent_wbs_id, w.wbs_short_name, w.wbs_name])

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    proj_name = schedule.projects[0].proj_short_name if schedule.projects else project_id
    filename = f"meridianiq-{proj_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════
# JSON / CSV Export (v1.2)
# ══════════════════════════════════════════════════════════


def _serialize_for_json(obj: Any) -> Any:
    """Recursively convert dataclass / datetime objects to JSON-safe dicts."""
    from datetime import date, datetime

    if obj is None or isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _serialize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_serialize_for_json(i) for i in obj]
    if hasattr(obj, "__dataclass_fields__"):
        return {k: _serialize_for_json(v) for k, v in asdict(obj).items()}
    if hasattr(obj, "model_dump"):
        return obj.model_dump()
    return str(obj)


def _build_export_data(
    schedule: ParsedSchedule,
    user_id: str | None,
) -> dict[str, Any]:
    """Build a comprehensive data bundle for JSON/CSV export.

    Runs DCMA, health score, and float analysis automatically.
    Returns a dict with all analysis results.
    """
    from src.analytics.dcma14 import DCMA14Analyzer
    from src.analytics.health_score import HealthScoreCalculator

    project_name = ""
    data_date = None
    if schedule.projects:
        proj = schedule.projects[0]
        project_name = proj.proj_short_name or ""
        data_date = proj.last_recalc_date or proj.sum_data_date

    dcma = DCMA14Analyzer(schedule)
    dcma_result = dcma.analyze()

    calc = HealthScoreCalculator(schedule)
    health = calc.calculate()

    # Activity summary
    activities_data = []
    for a in schedule.activities:
        activities_data.append(
            {
                "task_id": a.task_id,
                "task_code": a.task_code,
                "task_name": a.task_name,
                "task_type": a.task_type,
                "status_code": a.status_code,
                "total_float_hr": a.total_float_hr_cnt,
                "remaining_duration_hr": a.remain_drtn_hr_cnt,
                "original_duration_hr": a.target_drtn_hr_cnt,
                "physical_pct_complete": getattr(a, "phys_complete_pct", None),
                "actual_start": _serialize_for_json(a.act_start_date),
                "actual_finish": _serialize_for_json(a.act_end_date),
                "early_start": _serialize_for_json(a.early_start_date),
                "early_finish": _serialize_for_json(a.early_end_date),
                "late_start": _serialize_for_json(getattr(a, "late_start_date", None)),
                "late_finish": _serialize_for_json(getattr(a, "late_end_date", None)),
                "wbs_id": getattr(a, "wbs_id", None),
            }
        )

    # DCMA metrics
    dcma_metrics = []
    if hasattr(dcma_result, "metrics"):
        for m in dcma_result.metrics:
            dcma_metrics.append(
                {
                    "number": m.number,
                    "name": m.name,
                    "value": m.value,
                    "threshold": m.threshold,
                    "unit": m.unit,
                    "passed": m.passed,
                    "description": getattr(m, "description", ""),
                }
            )

    entropy = compute_float_entropy(schedule)

    return {
        "export_version": "1.0",
        "generator": "MeridianIQ",
        "project": {
            "name": project_name,
            "data_date": _serialize_for_json(data_date),
            "activity_count": len(schedule.activities),
            "relationship_count": len(schedule.relationships),
            "wbs_count": len(schedule.wbs_nodes),
            "calendar_count": len(schedule.calendars),
        },
        "health_score": {
            "overall": health.overall,
            "rating": health.rating,
            "trend": health.trend_arrow,
            "dcma_raw": health.dcma_raw,
            "float_raw": health.float_raw,
            "logic_raw": health.logic_raw,
            "trend_raw": health.trend_raw,
            "dcma_component": health.dcma_component,
            "float_component": health.float_component,
            "logic_component": health.logic_component,
            "trend_component": health.trend_component,
        },
        "dcma": {
            "overall_score": dcma_result.overall_score
            if hasattr(dcma_result, "overall_score")
            else 0,
            "passed_count": dcma_result.passed_count if hasattr(dcma_result, "passed_count") else 0,
            "failed_count": dcma_result.failed_count if hasattr(dcma_result, "failed_count") else 0,
            "metrics": dcma_metrics,
        },
        "float_entropy": {
            "entropy": entropy.entropy,
            "normalised_entropy": entropy.normalised_entropy,
            "max_entropy": entropy.max_entropy,
            "bucket_count": entropy.bucket_count,
            "distribution": entropy.distribution,
            "interpretation": entropy.interpretation,
        },
        "activities": activities_data,
        "relationships": [
            {
                "task_id": r.task_id,
                "pred_task_id": r.pred_task_id,
                "pred_type": r.pred_type,
                "lag_hr": r.lag_hr_cnt,
            }
            for r in schedule.relationships
        ],
    }


@app.get("/api/v1/projects/{project_id}/export/json")
def export_json(
    project_id: str,
    _user: object = Depends(optional_auth),
) -> StreamingResponse:
    """Export project schedule data and analysis results as JSON.

    Includes project metadata, DCMA 14-point assessment, health score,
    activities with dates/float, and relationships. Designed for
    researchers and external analysis tools.

    Args:
        project_id: The stored project identifier.

    Raises:
        HTTPException: If the project is not found.
    """
    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    data = _build_export_data(schedule, user_id)

    json_bytes = json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")

    proj_name = data["project"]["name"] or project_id
    filename = f"meridianiq-{proj_name}.json"

    return StreamingResponse(
        io.BytesIO(json_bytes),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/v1/projects/{project_id}/export/csv")
def export_csv(
    project_id: str,
    dataset: str = "activities",
    _user: object = Depends(optional_auth),
) -> StreamingResponse:
    """Export project data as CSV.

    Supports multiple datasets via the ``dataset`` query parameter:
    - ``activities`` (default): All activities with dates, float, status
    - ``dcma``: DCMA 14-point check results
    - ``relationships``: All predecessor relationships

    Args:
        project_id: The stored project identifier.
        dataset: Which dataset to export (activities, dcma, relationships).

    Raises:
        HTTPException: If the project is not found or dataset is invalid.
    """
    valid_datasets = {"activities", "dcma", "relationships"}
    if dataset not in valid_datasets:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid dataset '{dataset}'. Valid: {', '.join(sorted(valid_datasets))}",
        )

    store = get_store()
    user_id = _user["id"] if _user else None
    schedule = store.get(project_id, user_id=user_id)
    if schedule is None:
        raise HTTPException(status_code=404, detail="Project not found")

    data = _build_export_data(schedule, user_id)
    buf = io.StringIO()
    writer = csv.writer(buf)

    if dataset == "activities":
        headers = [
            "task_id",
            "task_code",
            "task_name",
            "task_type",
            "status_code",
            "total_float_hr",
            "remaining_duration_hr",
            "original_duration_hr",
            "physical_pct_complete",
            "actual_start",
            "actual_finish",
            "early_start",
            "early_finish",
            "late_start",
            "late_finish",
            "wbs_id",
        ]
        writer.writerow(headers)
        for a in data["activities"]:
            writer.writerow([a.get(h, "") for h in headers])

    elif dataset == "dcma":
        headers = ["number", "name", "value", "threshold", "unit", "passed", "description"]
        writer.writerow(headers)
        for m in data["dcma"]["metrics"]:
            writer.writerow([m.get(h, "") for h in headers])

    elif dataset == "relationships":
        headers = ["task_id", "pred_task_id", "pred_type", "lag_hr"]
        writer.writerow(headers)
        for r in data["relationships"]:
            writer.writerow([r.get(h, "") for h in headers])

    proj_name = data["project"]["name"] or project_id
    filename = f"meridianiq-{proj_name}-{dataset}.csv"

    csv_bytes = buf.getvalue().encode("utf-8")

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════
# IPS Reconciliation (v1.0)
# ══════════════════════════════════════════════════════════


@app.post("/api/v1/ips/reconcile")
def reconcile_ips(
    request_body: dict,
    _user: object = Depends(optional_auth),
) -> dict:
    """Run IPS reconciliation between a master schedule and sub-schedules.

    Per AACE RP 71R-12. Checks milestone alignment, date consistency,
    float consistency, and WBS alignment.

    Request body:
        master_project_id: str — the master schedule project ID
        sub_project_ids: list[str] — sub-schedule project IDs

    Returns:
        IPSReconciliationResult as dict.
    """
    from src.analytics.ips_reconciliation import IPSReconciler
    from dataclasses import asdict

    master_id = request_body.get("master_project_id")
    sub_ids = request_body.get("sub_project_ids", [])

    if not master_id or not sub_ids:
        raise HTTPException(
            status_code=400,
            detail="master_project_id and sub_project_ids are required",
        )

    store = get_store()
    user_id = _user["id"] if _user else None

    master = store.get(master_id, user_id=user_id)
    if master is None:
        raise HTTPException(status_code=404, detail=f"Master project {master_id} not found")

    subs = []
    for sid in sub_ids:
        sub = store.get(sid, user_id=user_id)
        if sub is None:
            raise HTTPException(status_code=404, detail=f"Sub-schedule {sid} not found")
        subs.append(sub)

    reconciler = IPSReconciler(master)
    result = reconciler.reconcile(subs)

    return asdict(result)


# ══════════════════════════════════════════════════════════
# Recovery Schedule Validation (v1.0)
# ══════════════════════════════════════════════════════════


@app.post("/api/v1/recovery/validate")
def validate_recovery(
    request_body: dict,
    _user: object = Depends(optional_auth),
) -> dict:
    """Validate a recovery schedule against the impacted schedule.

    Per AACE RP 29R-03 Section 4. Checks duration compression,
    scope changes, float consumption, and logic integrity.

    Request body:
        impacted_project_id: str — the impacted schedule
        recovery_project_id: str — the proposed recovery schedule
    """
    from src.analytics.recovery_validation import RecoveryValidator
    from dataclasses import asdict

    impacted_id = request_body.get("impacted_project_id")
    recovery_id = request_body.get("recovery_project_id")

    if not impacted_id or not recovery_id:
        raise HTTPException(
            status_code=400,
            detail="impacted_project_id and recovery_project_id are required",
        )

    store = get_store()
    user_id = _user["id"] if _user else None

    impacted = store.get(impacted_id, user_id=user_id)
    if impacted is None:
        raise HTTPException(status_code=404, detail=f"Impacted schedule {impacted_id} not found")

    recovery = store.get(recovery_id, user_id=user_id)
    if recovery is None:
        raise HTTPException(status_code=404, detail=f"Recovery schedule {recovery_id} not found")

    validator = RecoveryValidator(impacted, recovery)
    result = validator.validate()

    return asdict(result)


# ══════════════════════════════════════════════════════════
# API Keys (v1.2)
# ══════════════════════════════════════════════════════════


@app.post("/api/v1/api-keys")
def create_api_key(
    body: dict,
    _user: object = Depends(require_auth),
) -> dict:
    """Generate a new API key for programmatic access.

    The raw key is returned only once — store it securely.
    Subsequent requests use the key via ``X-API-Key`` header.

    Args:
        body: JSON with optional ``name`` field.

    Returns:
        Dict with ``key`` (raw), ``key_id``, ``name``, ``created_at``.
    """
    from src.api.auth import generate_api_key

    user_id = _user["id"]
    name = body.get("name", "default")
    result = generate_api_key(user_id, name)
    return result


@app.get("/api/v1/api-keys")
def list_api_keys_endpoint(
    _user: object = Depends(require_auth),
) -> dict:
    """List all API keys for the authenticated user.

    Does not return raw keys — only key_id, name, created_at.
    """
    from src.api.auth import list_api_keys

    user_id = _user["id"]
    keys = list_api_keys(user_id)
    return {"keys": keys}


@app.delete("/api/v1/api-keys/{key_id}")
def revoke_api_key_endpoint(
    key_id: str,
    _user: object = Depends(require_auth),
) -> dict:
    """Revoke an API key.

    Args:
        key_id: The key identifier to revoke.

    Returns:
        Dict with ``revoked`` boolean.
    """
    from src.api.auth import revoke_api_key

    user_id = _user["id"]
    revoked = revoke_api_key(user_id, key_id)
    if not revoked:
        raise HTTPException(status_code=404, detail="API key not found")
    return {"revoked": True, "key_id": key_id}
